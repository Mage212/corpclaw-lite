"""excel_inspect -- lightweight file structure inspection for the main agent.

Returns metadata about tabular files (sheets, dimensions, merged cells,
color groups, sample values) without loading data into memory.  The main
agent calls this before ``dispatch_subagent`` to decide which subagent
and toolset is appropriate for the task.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, cast

from corpclaw_lite.extensions.tools.base import RiskLevel, Tool, ToolParam
from corpclaw_lite.extensions.tools.builtin.files import resolve_and_validate_path
from corpclaw_lite.extensions.tools.builtin.table_query import detect_csv_encoding
from corpclaw_lite.utils.async_helpers import run_in_thread

__all__ = ["ExcelInspectTool"]

_MAX_PREVIEW_ROWS = 5
_MAX_MERGED_SHOW = 15
_MAX_FULL_XLSX_BYTES = 20 * 1024 * 1024
_MAX_CSV_COUNT_BYTES = 20 * 1024 * 1024
_MAX_CSV_COUNT_ROWS = 10_000


def _format_size(size_bytes: int) -> str:
    if size_bytes < 1024:
        return f"{size_bytes} B"
    if size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    return f"{size_bytes / (1024 * 1024):.1f} MB"


def _inspect_xlsx(path: Path, detail: str) -> str:
    import openpyxl

    size = path.stat().st_size
    if detail == "full" and size > _MAX_FULL_XLSX_BYTES:
        return (
            f"File: {path.name} ({_format_size(size)})\n"
            f"Error: XLSX file too large for full inspection ({_format_size(size)}). "
            f"Limit: {_format_size(_MAX_FULL_XLSX_BYTES)}. Use detail='summary'."
        )

    wb = openpyxl.load_workbook(str(path), read_only=(detail == "summary"), data_only=False)
    try:
        lines: list[str] = [f"File: {path.name} ({_format_size(path.stat().st_size)})"]
        lines.append(f"Sheets: {len(wb.sheetnames)}")

        for idx, name in enumerate(wb.sheetnames, 1):
            ws = wb[name]
            rows = ws.max_row or 0
            cols = ws.max_column or 0
            merged_ranges = getattr(getattr(ws, "merged_cells", None), "ranges", [])
            merged_count = len(merged_ranges)

            lines.append(
                f'\n--- Sheet {idx}: "{name}" '
                f"({rows} rows x {cols} cols, {merged_count} merged) ---"
            )

            if detail == "summary":
                continue

            # Merged ranges with values
            if ws.merged_cells.ranges:
                lines.append(f"\n  Merged ranges (top {_MAX_MERGED_SHOW}):")
                for mr in list(ws.merged_cells.ranges)[:_MAX_MERGED_SHOW]:
                    cell = ws.cell(row=mr.min_row, column=mr.min_col)
                    val = repr(cell.value)[:60] if cell.value is not None else "(empty)"
                    lines.append(f"    {mr} = {val}")
                if len(ws.merged_cells.ranges) > _MAX_MERGED_SHOW:
                    extra = len(ws.merged_cells.ranges) - _MAX_MERGED_SHOW
                    lines.append(f"    ... and {extra} more")

            # Color groups — compact summary
            color_counts: dict[str, int] = {}
            for row_idx in range(1, min(rows + 1, 100)):
                for col_idx in range(1, cols + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    fill = cell.fill
                    if not fill.patternType or fill.patternType == "none":  # type: ignore[unknownMemberType]
                        continue
                    fg = fill.fgColor  # type: ignore[unknownMemberType]
                    if fg is None or fg.rgb is None or fg.rgb in ("00000000", "0"):  # type: ignore[unknownMemberType]
                        continue
                    try:
                        rgb = str(fg.rgb)  # type: ignore[unknownMemberType]
                    except Exception:
                        continue
                    display_rgb = rgb[2:] if len(rgb) == 8 and rgb.startswith("FF") else rgb
                    color_counts[display_rgb] = color_counts.get(display_rgb, 0) + 1

            if color_counts:
                sorted_colors = sorted(color_counts.items(), key=lambda x: -x[1])
                color_parts = [f"#{rgb}: {cnt} cells" for rgb, cnt in sorted_colors]
                summary = ", ".join(color_parts)
                lines.append(f"\n  Color groups: {len(sorted_colors)} colors ({summary})")

            # Preview: first rows
            if rows > 0:
                lines.append(f"\n  Preview (first {_MAX_PREVIEW_ROWS} rows):")
                for row_idx in range(1, min(rows + 1, _MAX_PREVIEW_ROWS + 1)):
                    cells_info: list[str] = []
                    for col_idx in range(1, min(cols + 1, 12)):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value is not None:
                            val_str = str(cell.value)[:30]
                            cells_info.append(f"{cell.coordinate}={val_str}")
                    if cells_info:
                        lines.append(f"    Row {row_idx}: {'  |  '.join(cells_info)}")
                    else:
                        lines.append(f"    Row {row_idx}: (empty)")

        return "\n".join(lines)
    finally:
        wb.close()


def _inspect_csv(path: Path) -> str:
    enc = detect_csv_encoding(path)
    size = path.stat().st_size
    lines: list[str] = [f"File: {path.name} ({_format_size(size)})"]
    lines.append(f"Format: CSV (encoding: {enc})")

    with open(path, newline="", encoding=enc) as f:
        reader = csv.reader(f)
        headers = next(reader, None)
        if headers:
            lines.append(f"Columns ({len(headers)}): {', '.join(str(h) for h in headers[:20])}")

        # Count rows and collect preview.  Large CSVs are sampled to avoid
        # spending unbounded time just to count rows for a lightweight inspect.
        row_count = 0
        preview: list[list[str]] = []
        for row in reader:
            row_count += 1
            if len(preview) < _MAX_PREVIEW_ROWS:
                preview.append(row)
            if size > _MAX_CSV_COUNT_BYTES and row_count >= _MAX_CSV_COUNT_ROWS:
                break
        if size > _MAX_CSV_COUNT_BYTES and row_count >= _MAX_CSV_COUNT_ROWS:
            lines.append(
                f"Data rows: \u2265{row_count} (large file; exact count skipped for performance)"
            )
        else:
            lines.append(f"Data rows: {row_count}")

        if preview:
            lines.append(f"\nPreview (first {len(preview)} rows):")
            for i, row in enumerate(preview, 1):
                cells = [str(v)[:30] for v in row[:12]]
                lines.append(f"  Row {i}: {' | '.join(cells)}")

    return "\n".join(lines)


_MAX_JSON_SIZE = 50 * 1024 * 1024  # 50 MB


def _inspect_json(path: Path) -> str:
    import json

    size = path.stat().st_size
    if size > _MAX_JSON_SIZE:
        return (
            f"File: {path.name} ({_format_size(size)})\n"
            f"Error: JSON file too large for inspection ({_format_size(size)}). "
            f"Limit: {_format_size(_MAX_JSON_SIZE)}."
        )

    data: Any = json.loads(path.read_text(encoding="utf-8"))
    lines: list[str] = [f"File: {path.name} ({_format_size(path.stat().st_size)})"]
    lines.append("Format: JSON")

    if isinstance(data, list):
        items: list[Any] = data  # type: ignore[assignment]
        lines.append(f"Type: array of {len(items)} items")
        if items and isinstance(items[0], dict):
            first_item = cast(dict[str, Any], items[0])
            cols = list(first_item.keys())
            col_names = [str(c) for c in cols[:20]]
            lines.append(f"Columns ({len(cols)}): {', '.join(col_names)}")
            for i, raw_item in enumerate(items[:_MAX_PREVIEW_ROWS], 1):
                typed_item = cast(dict[str, Any], raw_item)
                pairs = list(typed_item.items())[:8]
                vals = [f"{k}={str(v)[:25]}" for k, v in pairs]
                lines.append(f"  Item {i}: {' | '.join(vals)}")
    elif isinstance(data, dict):
        obj = cast(dict[str, Any], data)
        lines.append(f"Type: object with {len(obj)} keys")
        keys = list(obj.keys())[:15]
        lines.append(f"Keys: {', '.join(str(k) for k in keys)}")
    else:
        lines.append(f"Type: {type(data).__name__}")

    return "\n".join(lines)


def _inspect_parquet(path: Path) -> str:
    try:
        import duckdb
    except ImportError:
        return f"File: {path.name}\nFormat: Parquet (duckdb not available for inspection)"

    conn = duckdb.connect(":memory:")
    try:
        p = str(path).replace("'", "''")
        conn.execute(f"CREATE TABLE _insp AS SELECT * FROM read_parquet('{p}')")
        result = conn.execute("SELECT COUNT(*) FROM _insp").fetchone()
        row_count = result[0] if result else 0

        cols_result = conn.execute(
            "SELECT column_name, column_type "
            "FROM information_schema.columns "
            "WHERE table_name='_insp'"
        ).fetchall()
        col_info = [f"{name} ({typ})" for name, typ in cols_result]

        lines = [f"File: {path.name} ({_format_size(path.stat().st_size)})"]
        lines.append("Format: Parquet")
        lines.append(f"Rows: {row_count}")
        lines.append(f"Columns ({len(col_info)}): {', '.join(col_info[:15])}")

        # Preview
        preview = conn.execute("SELECT * FROM _insp LIMIT 3").fetchall()
        if preview:
            lines.append(f"\nPreview (first {len(preview)} rows):")
            for i, row in enumerate(preview, 1):
                vals = [str(v)[:30] for v in row]
                lines.append(f"  Row {i}: {' | '.join(vals)}")

        return "\n".join(lines)
    finally:
        conn.close()


def _do_inspect(path: Path, detail: str) -> str:
    ext = path.suffix.lower()
    if ext == ".xlsx":
        return _inspect_xlsx(path, detail)
    if ext == ".csv":
        return _inspect_csv(path)
    if ext in (".json", ".jsonl", ".ndjson"):
        return _inspect_json(path)
    if ext == ".parquet":
        return _inspect_parquet(path)
    return f"Error: Unsupported file format '{ext}'. Use .xlsx, .csv, .json, or .parquet."


class ExcelInspectTool(Tool):
    """Lightweight file structure inspection for the main agent."""

    name = "excel_inspect"
    description = (
        "Quickly inspect a tabular file's structure: sheets, dimensions, "
        "merged cells, color groups, and a small preview. For detailed cell "
        "contents, reading specific ranges, or data analysis, dispatch a "
        "subagent (data-agent or document-agent) — they have excel_workbook "
        "and table_query tools."
    )
    params = [
        ToolParam(
            name="path",
            type="string",
            description="Path to file (.xlsx, .csv, .json, .parquet)",
        ),
        ToolParam(
            name="detail",
            type="string",
            description=(
                "Detail level: summary (sheets+dimensions) or full (+merged, colors, samples)"
            ),
            enum=["summary", "full"],
            required=False,
        ),
    ]
    risk_level = RiskLevel.LOW
    parallel_safe = True

    async def execute(self, **kwargs: Any) -> str:
        path_str = kwargs.get("path", "")
        if not path_str:
            return "Error: 'path' is required."

        try:
            resolved = resolve_and_validate_path(path_str)
        except PermissionError as e:
            return f"Error: {e}"

        if not resolved.is_file():
            return f"Error: File not found: {path_str}"

        detail = kwargs.get("detail", "summary")
        return await run_in_thread(_do_inspect, resolved, detail)
