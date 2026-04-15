"""normalize_excel — Excel file normalization tool.

Restores original logic from ord_report_unification_glm/tools/normalize_format.py:
- Removes invisible/problematic characters
- Strict cell type enforcement (INN as text, dates as DD.MM.YYYY, numbers rounded)
- Preserves original headers (no snake_case, no renaming)
- Creates new formatted workbook with borders, fonts, auto-width, freeze panes
- Removes completely empty rows
- Does NOT remove duplicates or modify headers
"""

from __future__ import annotations

import math
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from corpclaw_lite.extensions.tools.base import RiskLevel, Tool, ToolParam
from corpclaw_lite.extensions.tools.builtin.files import resolve_and_validate_path
from corpclaw_lite.utils.async_helpers import run_in_thread

__all__ = [
    "NormalizeExcelTool",
]

INVISIBLE_CHARS: dict[str, str] = {
    "\u00ad": "",
    "\u200b": "",
    "\u200c": "",
    "\u200d": "",
    "\ufeff": "",
    "\u200e": "",
    "\u200f": "",
    "\u2028": " ",
    "\u2029": " ",
    "\u2060": "",
    "\u00a0": " ",
}


def _clean_chars(text: str) -> str:
    """Remove invisible and problematic characters."""
    for char, repl in INVISIBLE_CHARS.items():
        text = text.replace(char, repl)
    text = "".join(c for c in text if ord(c) >= 32 or c in "\n\t")
    return text.strip()


def _detect_column_type(header: str) -> str:
    """Detect column type by header keywords.

    Order matches original: numeric (инвентарь, сумма) → inn → date → text.
    """
    low = header.lower().strip()
    if "инвентарь" in low or "сумма" in low:
        return "numeric"
    if "инн" in low:
        return "inn"
    if "дата" in low:
        return "date"
    return "text"


def _fix_value(val: Any, col_type: str) -> Any:
    """Normalize a single cell value based on column type.

    Faithfully reproduces the logic from the original normalize_format.py
    with two improvements: leading-zero INN restoration and comma-as-decimal.
    """
    if val is None:
        return None

    if isinstance(val, datetime):
        if col_type == "date":
            return val.strftime("%d.%m.%Y")
        return None

    try:
        import pandas as pd  # type: ignore[import-untyped]

        if isinstance(val, pd.Timestamp):  # type: ignore[union-attr]
            if col_type == "date":
                return val.strftime("%d.%m.%Y")
            return None
    except ImportError:
        pass

    if isinstance(val, str):
        val = _clean_chars(val)
        if not val:
            return None

        if col_type == "numeric":
            try:
                s = val.replace(",", ".").strip()
                if "." in s:
                    return round(float(s), 2)
                return int(s)
            except ValueError:
                return val

        if col_type == "inn":
            s = val.strip()
            if "e+" in s.lower() or "e-" in s.lower():
                try:
                    return str(int(float(s)))
                except (ValueError, OverflowError):
                    pass
            return s

        return val

    # Float values
    if isinstance(val, float):
        if math.isnan(val) or math.isinf(val):
            return None

        if col_type == "inn":
            s = str(int(val))
            if len(s) in (9, 11):
                s = "0" + s
            return s

        if col_type == "numeric":
            return round(val, 2)

        if col_type == "date" and 40000 < val < 60000:
            dt = datetime(1899, 12, 30) + timedelta(days=int(val))
            return dt.strftime("%d.%m.%Y")

        # Text columns: float 123.0 → "123"
        if val.is_integer():
            return str(int(val))
        return str(val)

    # Int values (bool excluded — True/False should not be treated as numbers)
    if isinstance(val, int) and not isinstance(val, bool):
        if col_type == "inn":
            s = str(val)
            if len(s) in (9, 11):
                s = "0" + s
            return s

        if col_type == "numeric":
            return val

        if col_type == "date" and 40000 < val < 60000:
            dt = datetime(1899, 12, 30) + timedelta(days=val)
            return dt.strftime("%d.%m.%Y")

        return str(val)

    return val


def _is_empty_row(values: list[Any]) -> bool:
    """Check if all values in a row are empty."""
    return all(v is None or (isinstance(v, str) and v.strip() == "") for v in values)


def _create_normalized_workbook(
    headers: list[str],
    rows: list[list[Any]],
    col_types: dict[int, str],
) -> Any:
    """Create a new formatted Workbook — matches original create_normalized_workbook."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Данные"

    header_font = Font(name="Calibri", size=11, bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    col_formats: dict[int, str] = {}
    for col_idx, col_type in col_types.items():
        col_formats[col_idx] = "General" if col_type == "numeric" else "@"

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.number_format = "@"

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            cell.number_format = col_formats.get(col_idx, "@")

    for col_idx in range(1, len(headers) + 1):
        max_length = len(str(headers[col_idx - 1]))
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = float(min(max_length + 2, 50))

    ws.freeze_panes = "A2"
    return wb


class NormalizeExcelTool(Tool):
    """Normalize an Excel (.xlsx) file: fix types, clean invisible chars, apply formatting.

    Creates a new workbook with proper cell formats — INN as text, dates as
    DD.MM.YYYY, numbers rounded to 2 decimals.  Headers are preserved as-is.
    """

    name = "normalize_excel"
    description = (
        "Normalize an Excel (.xlsx) file: fix INN (remove scientific notation, "
        "restore leading zeros), convert dates to DD.MM.YYYY, round numbers, "
        "remove invisible characters, apply standard formatting."
    )
    params = [
        ToolParam(name="path", type="string", description="Path to .xlsx file"),
        ToolParam(
            name="output_path",
            type="string",
            description="Output path (default: {name}_normalized.xlsx)",
            required=False,
        ),
        ToolParam(
            name="remove_empty_rows",
            type="boolean",
            description="Remove completely empty rows (default true)",
            required=False,
        ),
    ]
    risk_level = RiskLevel.MEDIUM

    async def execute(self, **kwargs: Any) -> str:
        path_str = kwargs.get("path")
        if not path_str or not isinstance(path_str, str):
            return "Error: 'path' parameter is required."

        try:
            resolved = resolve_and_validate_path(path_str)
        except PermissionError as e:
            return f"Error: {e}"

        if not resolved.exists():
            return f"Error: File '{path_str}' does not exist."

        if resolved.suffix.lower() != ".xlsx":
            return "Error: Only .xlsx files are supported."

        output_str = kwargs.get("output_path")
        if output_str and isinstance(output_str, str):
            try:
                output_path = resolve_and_validate_path(output_str)
            except PermissionError as e:
                return f"Error: {e}"
        else:
            output_path = resolved.parent / f"{resolved.stem}_normalized.xlsx"

        do_empty = kwargs.get("remove_empty_rows", True)

        try:
            return await run_in_thread(self._process_file, resolved, output_path, do_empty)
        except Exception as e:
            return f"Error: {e}"

    def _process_file(self, resolved: Path, output_path: Path, do_empty: bool) -> str:
        """Synchronous Excel processing — runs in thread pool via anyio."""
        try:
            import openpyxl
        except ImportError:
            return "Error: openpyxl is not installed."

        try:
            wb_src = openpyxl.load_workbook(str(resolved), data_only=True)
        except Exception as e:
            return f"Error reading Excel file: {e}"

        ws_src = wb_src.active
        if ws_src is None:
            return "Error: Workbook has no active sheet."

        total_rows = ws_src.max_row or 0
        total_cols = ws_src.max_column or 0

        if total_rows == 0:
            wb_src.close()
            return "Error: File is empty."

        headers: list[str] = []
        col_types: dict[int, str] = {}
        for col in range(1, total_cols + 1):
            raw = ws_src.cell(row=1, column=col).value
            if isinstance(raw, str):
                header = _clean_chars(raw)
            else:
                header = str(raw) if raw is not None else ""
            headers.append(header)
            col_types[col] = _detect_column_type(header)

        # Trim trailing empty headers caused by max_column including
        # formatted-but-empty columns (borders, number format, etc.)
        while headers and not headers[-1].strip():
            headers.pop()
        total_cols = len(headers)

        data_rows: list[list[Any]] = []
        values_fixed = 0
        empty_removed = 0

        for row_idx in range(2, total_rows + 1):
            row_values: list[Any] = []
            for col in range(1, total_cols + 1):
                val = ws_src.cell(row=row_idx, column=col).value
                fixed = _fix_value(val, col_types.get(col, "text"))
                if fixed != val:
                    values_fixed += 1
                row_values.append(fixed)

            if do_empty and _is_empty_row(row_values):
                empty_removed += 1
                continue

            data_rows.append(row_values)

        wb_src.close()

        wb_out = _create_normalized_workbook(headers, data_rows, col_types)

        try:
            wb_out.save(str(output_path))
        except Exception as e:
            return f"Error saving file: {e}"

        total_data = total_rows - 1
        remaining = len(data_rows)

        parts = [f"Normalized '{resolved.name}' → '{output_path.name}'"]
        parts.append(f"Rows processed: {total_data}, remaining: {remaining}")
        if values_fixed:
            parts.append(f"Values cleaned/fixed: {values_fixed}")
        if empty_removed:
            parts.append(f"Empty rows removed: {empty_removed}")
        parts.append(f"Output filename: {output_path.name}")

        return "\n".join(parts)
