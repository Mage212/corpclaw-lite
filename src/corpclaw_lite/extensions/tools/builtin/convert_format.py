"""convert_format — convert tabular data between CSV, XLSX, JSON, and Markdown.

Reads a source file, detects its format by extension, loads into an
intermediate representation (list of dicts), then writes to the target format.
"""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from corpclaw_lite.extensions.tools.base import RiskLevel, Tool, ToolParam
from corpclaw_lite.extensions.tools.builtin.files import resolve_and_validate_path
from corpclaw_lite.extensions.tools.builtin.table_query import detect_csv_encoding
from corpclaw_lite.utils.async_helpers import run_in_thread

__all__ = ["ConvertFormatTool"]

_SUPPORTED_INPUT = {".csv", ".xlsx", ".json", ".md", ".markdown"}
_SUPPORTED_OUTPUT = {"csv", "xlsx", "json", "markdown"}
_MAX_INPUT_BYTES = 50 * 1024 * 1024


# --- Loaders ---


def _load_csv(path: Path) -> list[dict[str, str]]:
    encoding = detect_csv_encoding(path)
    with open(path, newline="", encoding=encoding) as f:
        reader = csv.DictReader(f)
        return list(reader)  # pyright: ignore[reportUnknownVariableType]


def _load_json(path: Path) -> list[dict[str, Any]]:
    data: Any = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, list):
        return data  # pyright: ignore[reportUnknownVariableType]
    if isinstance(data, dict):
        # Single object → wrap in list.
        return [data]
    raise ValueError("JSON must be an array of objects or a single object.")


def _load_xlsx(path: Path) -> list[dict[str, Any]]:
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            return []
        rows_iter = ws.iter_rows(values_only=True)
        headers: list[str] | None = None
        data: list[dict[str, Any]] = []
        for row in rows_iter:
            if headers is None:
                headers = [
                    str(c).strip() if c is not None else f"col{i}" for i, c in enumerate(row)
                ]
                continue
            data.append(dict(zip(headers, row, strict=False)))
        return data
    finally:
        wb.close()


def _load_markdown(path: Path) -> list[dict[str, str]]:
    """Parse a Markdown pipe-delimited table."""
    lines = path.read_text(encoding="utf-8").splitlines()
    # Find table lines (start and end with '|').
    table_lines = [
        ln.strip() for ln in lines if ln.strip().startswith("|") and ln.strip().endswith("|")
    ]
    if not table_lines:
        raise ValueError("No Markdown table found in file.")

    def _split_row(line: str) -> list[str]:
        return [cell.strip() for cell in line.strip("|").split("|")]

    headers = _split_row(table_lines[0])
    # Skip separator row (contains only -, :, |).
    data_lines: list[dict[str, str]] = []
    for line in table_lines[1:]:
        cells = _split_row(line)
        if all(set(c) <= {"-", ":"} for c in cells):
            continue  # separator row
        data_lines.append(dict(zip(headers, cells, strict=False)))  # pyright: ignore[reportUnknownMemberType]
    return data_lines


# --- Writers ---


def _write_csv(data: list[dict[str, Any]], path: Path) -> None:
    if not data:
        path.write_text("", encoding="utf-8-sig")
        return
    fieldnames = list(data[0].keys())
    # utf-8-sig writes BOM so Excel correctly recognises Cyrillic.
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in data:
            writer.writerow({k: (v if v is not None else "") for k, v in row.items()})


def _write_json(data: list[dict[str, Any]], path: Path) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2, default=str), encoding="utf-8")


def _write_xlsx(data: list[dict[str, Any]], path: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None  # openpyxl Workbook always has an active sheet
    if not data:
        wb.save(str(path))
        return
    headers = list(data[0].keys())
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, row in enumerate(data, 2):
        for col, h in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col, value=row.get(h))
    wb.save(str(path))


def _write_markdown(data: list[dict[str, Any]], path: Path) -> None:
    if not data:
        path.write_text("", encoding="utf-8")
        return
    headers = list(data[0].keys())
    col_widths = [len(h) for h in headers]
    for row in data:
        for i, h in enumerate(headers):
            col_widths[i] = max(col_widths[i], len(str(row.get(h, ""))))

    lines: list[str] = []
    lines.append("| " + " | ".join(h.ljust(col_widths[i]) for i, h in enumerate(headers)) + " |")
    lines.append("| " + " | ".join("-" * col_widths[i] for i in range(len(headers))) + " |")
    lines.extend(
        "| "
        + " | ".join(str(row.get(h, "")).ljust(col_widths[i]) for i, h in enumerate(headers))
        + " |"
        for row in data
    )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _do_convert(input_path: Path, output_format: str, output_path: Path | None) -> str:
    ext = input_path.suffix.lower()

    # Load.
    if ext == ".csv":
        data = _load_csv(input_path)
    elif ext == ".json":
        data = _load_json(input_path)
    elif ext == ".xlsx":
        data = _load_xlsx(input_path)
    elif ext in (".md", ".markdown"):
        data = _load_markdown(input_path)
    else:
        supported = ", ".join(sorted(_SUPPORTED_INPUT))
        return f"Error: Unsupported input format '{ext}'. Supported: {supported}"

    if not data:
        return "Error: No data found in input file."

    # Determine output path.
    if output_path is None:
        if output_format == "markdown":
            output_path = input_path.with_suffix(".md")
        else:
            output_path = input_path.with_suffix(f".{output_format}")

    # Write.
    if output_format == "csv":
        _write_csv(data, output_path)
    elif output_format == "json":
        _write_json(data, output_path)
    elif output_format == "xlsx":
        _write_xlsx(data, output_path)
    elif output_format == "markdown":
        _write_markdown(data, output_path)
    else:
        return f"Error: Unsupported output format '{output_format}'."

    return (
        f"Converted {input_path.name} ({ext}) → {output_path.name} ({output_format})\n"
        f"Rows: {len(data)}, Columns: {len(data[0].keys()) if data else 0}"
    )


class ConvertFormatTool(Tool):
    """Convert tabular data between CSV, XLSX, JSON, and Markdown formats."""

    name = "convert_format"
    description = (
        "Convert tabular data between CSV, XLSX, JSON, and Markdown table formats. "
        "Input format is detected from file extension."
    )
    params = [
        ToolParam(
            name="input_path",
            type="string",
            description="Path to source file (CSV, XLSX, JSON, or Markdown)",
        ),
        ToolParam(
            name="output_format",
            type="string",
            description="Target format: csv, xlsx, json, or markdown",
            enum=["csv", "xlsx", "json", "markdown"],
        ),
        ToolParam(
            name="output_path",
            type="string",
            description="Custom output file path (auto-generated if omitted)",
            required=False,
        ),
    ]
    risk_level = RiskLevel.MEDIUM

    async def execute(self, **kwargs: Any) -> str:
        input_str = kwargs.get("input_path", "")
        output_format = kwargs.get("output_format", "")
        output_str = kwargs.get("output_path")

        if not input_str:
            return "Error: 'input_path' is required."
        if not output_format:
            return "Error: 'output_format' is required."
        if output_format not in _SUPPORTED_OUTPUT:
            return f"Error: Unsupported output format '{output_format}'."

        try:
            resolved_input = resolve_and_validate_path(input_str)
        except PermissionError as e:
            return f"Error: {e}"

        if not resolved_input.is_file():
            return f"Error: File not found: {input_str}"
        if resolved_input.stat().st_size > _MAX_INPUT_BYTES:
            size_mb = resolved_input.stat().st_size / (1024 * 1024)
            limit_mb = _MAX_INPUT_BYTES / (1024 * 1024)
            return (
                f"Error: Input file too large for convert_format "
                f"({size_mb:.1f} MB, limit {limit_mb:.1f} MB)."
            )

        output_path: Path | None = None
        if output_str:
            try:
                output_path = resolve_and_validate_path(output_str)
            except PermissionError as e:
                return f"Error: {e}"

        return await run_in_thread(_do_convert, resolved_input, output_format, output_path)
