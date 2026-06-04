"""table_query — run SQL queries on tabular data files (CSV, XLSX, JSON).

Loads the file into an in-memory DuckDB table named ``data`` and executes
the user-provided SQL query against it.  Results are returned as a
pipe-delimited text table and optionally saved to CSV.
"""

from __future__ import annotations

import csv
import datetime as dt
import re
from pathlib import Path
from typing import Any

from corpclaw_lite.extensions.tools.base import RiskLevel, Tool, ToolParam
from corpclaw_lite.extensions.tools.builtin.files import resolve_and_validate_path
from corpclaw_lite.utils.async_helpers import run_in_thread

__all__ = [
    "TableQueryTool",
    "load_xlsx_as_dicts",
    "load_all_xlsx_sheets",
    "init_duckdb_with_xlsx",
    "detect_csv_encoding",
    "reencode_csv_to_utf8",
    "_sanitize_table_name",
]

_MAX_RESULT_ROWS = 10_000
_MAX_RESULT_CHARS = 50_000
_SQL_READ_ONLY_ERROR = (
    "Error: table_query only supports read-only SELECT/WITH queries against loaded tables. "
    "Use the 'path' parameter to load files; file-reading SQL functions and DDL/DML are blocked."
)
_BLOCKED_SQL_PATTERNS = [
    r"\bATTACH\b",
    r"\bCOPY\b",
    r"\bCREATE\b",
    r"\bDELETE\b",
    r"\bDETACH\b",
    r"\bDROP\b",
    r"\bEXPORT\b",
    r"\bIMPORT\b",
    r"\bINSERT\b",
    r"\bINSTALL\b",
    r"\bLOAD\b",
    r"\bTRUNCATE\b",
    r"\bUPDATE\b",
    r"\bread_csv\s*\(",
    r"\bread_csv_auto\s*\(",
    r"\bread_json\s*\(",
    r"\bread_json_auto\s*\(",
    r"\bread_ndjson\s*\(",
    r"\bread_parquet\s*\(",
    r"\bread_text\s*\(",
    r"\bread_blob\s*\(",
    r"\bglob\s*\(",
    r"\bcsv_scan\s*\(",
    r"\bjson_scan\s*\(",
    r"\bparquet_scan\s*\(",
    r"\b[a-zA-Z_][a-zA-Z0-9_]*_scan\s*\(",
]


def detect_csv_encoding(path: Path) -> str:
    """Detect CSV file encoding.

    Priority: BOM → UTF-8 try → Windows-1251 fallback.
    Returns Python encoding names (e.g. ``utf-8-sig`` for BOM).
    Use :func:`duckdb_encoding` to convert for DuckDB.
    """
    raw = path.read_bytes()[:64]
    if raw[:3] == b"\xef\xbb\xbf":
        return "utf-8-sig"
    try:
        path.read_text(encoding="utf-8")
        return "utf-8"
    except UnicodeDecodeError:
        return "cp1251"


def reencode_csv_to_utf8(path: Path, encoding: str) -> Path:
    """Re-encode a non-UTF-8 CSV to a temporary UTF-8 file for DuckDB.

    DuckDB's ICU encoding support is incomplete (no windows-1251, etc.),
    so we re-encode via Python before loading.
    """
    import tempfile

    raw = path.read_bytes()
    text = raw.decode(encoding)
    # Strip BOM if present — DuckDB handles it but let's be safe.
    if text.startswith("\ufeff"):
        text = text[1:]

    with tempfile.NamedTemporaryFile(
        mode="w", suffix=".csv", prefix="cc_utf8_", encoding="utf-8", delete=False
    ) as tmp_file:
        tmp_file.write(text)
        return Path(tmp_file.name)


def load_xlsx_as_dicts(
    path: Path,
    sheet_name: str | None = None,
) -> list[dict[str, Any]]:
    """Load an XLSX file into a list of dicts using openpyxl.

    If *sheet_name* is given, load that specific sheet; otherwise use the
    active (first) sheet.
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    try:
        if sheet_name is not None:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
            ws = wb[sheet_name]
        else:
            ws = wb.active
        if ws is None:
            return []
        rows_iter = ws.iter_rows(values_only=True)
        headers: list[str] | None = None
        data: list[dict[str, Any]] = []
        for row in rows_iter:
            if headers is None:
                headers = [
                    str(c).strip() if c is not None else f"col{i}" for i, c in enumerate(row)
                ]
                continue
            data.append(dict(zip(headers, row, strict=False)))
        return data
    finally:
        wb.close()


def load_all_xlsx_sheets(path: Path) -> dict[str, list[dict[str, Any]]]:
    """Load all sheets from an XLSX file.  Returns {sheet_name: rows}."""
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    try:
        result: dict[str, list[dict[str, Any]]] = {}
        for name in wb.sheetnames:
            ws = wb[name]
            rows_iter = ws.iter_rows(values_only=True)
            headers: list[str] | None = None
            data: list[dict[str, Any]] = []
            for row in rows_iter:
                if headers is None:
                    headers = [
                        str(c).strip() if c is not None else f"col{i}" for i, c in enumerate(row)
                    ]
                    continue
                data.append(dict(zip(headers, row, strict=False)))
            result[name] = data
        return result
    finally:
        wb.close()


def _init_duckdb_table(
    conn: Any,
    table_name: str,
    rows: list[dict[str, Any]],
) -> None:
    """Create a DuckDB table from row dicts with the given table name."""
    if not rows:
        return
    columns = list(rows[0].keys())
    safe_cols = [c.replace('"', '""') for c in columns]
    col_defs = ", ".join(
        f'"{c}" {_infer_duckdb_type([row.get(col) for row in rows])}'
        for c, col in zip(safe_cols, columns, strict=False)
    )
    conn.execute(f'CREATE TABLE "{table_name}" ({col_defs})')
    conn.executemany(
        f'INSERT INTO "{table_name}" VALUES ({", ".join("?" for _ in columns)})',
        [tuple(row.get(col) for col in columns) for row in rows],
    )


def _infer_duckdb_type(values: list[Any]) -> str:
    """Infer a conservative DuckDB type for values loaded from XLSX."""
    non_null = [v for v in values if v is not None]
    if not non_null:
        return "VARCHAR"
    if all(isinstance(v, bool) for v in non_null):
        return "BOOLEAN"
    if all(isinstance(v, int) and not isinstance(v, bool) for v in non_null):
        return "BIGINT"
    if all(isinstance(v, int | float) and not isinstance(v, bool) for v in non_null):
        return "DOUBLE"
    if all(isinstance(v, dt.datetime) for v in non_null):
        return "TIMESTAMP"
    if all(isinstance(v, dt.date) and not isinstance(v, dt.datetime) for v in non_null):
        return "DATE"
    return "VARCHAR"


def init_duckdb_with_xlsx(path: Path, rows: list[dict[str, Any]], conn: Any) -> None:
    """Create table ``data`` from XLSX rows.  Backward-compatible wrapper."""
    _init_duckdb_table(conn, "data", rows)


def _sanitize_table_name(sheet_name: str) -> str:
    """Make a sheet name safe as a DuckDB table identifier."""
    name = sheet_name.strip()
    name = re.sub(r"[^a-zA-Z0-9_]", "_", name)
    name = re.sub(r"_+", "_", name).strip("_")
    if not name:
        name = f"sheet_{abs(hash(sheet_name)) % 10000}"
    elif name[0].isdigit():
        name = f"sheet_{name}"
    return name.lower()


def _format_results(columns: list[str], rows: list[tuple[Any, ...]]) -> str:
    """Format query results as a readable text table."""
    if not rows:
        return "Query returned 0 rows."

    # Truncate rows if needed.
    truncated = len(rows) > _MAX_RESULT_ROWS
    if truncated:
        rows = rows[:_MAX_RESULT_ROWS]

    # Compute column widths.
    col_widths = [len(c) for c in columns]
    for row in rows:
        for i, val in enumerate(row):
            col_widths[i] = max(col_widths[i], len(str(val)))

    # Build separator.
    sep = "+" + "+".join("-" * (w + 2) for w in col_widths) + "+"
    header = "|" + "|".join(f" {c:<{col_widths[i]}} " for i, c in enumerate(columns)) + "|"

    lines = [sep, header, sep]
    for row in rows:
        line = "|" + "|".join(f" {str(val):<{col_widths[i]}} " for i, val in enumerate(row)) + "|"
        lines.append(line)
    lines.append(sep)

    result = "\n".join(lines)
    if truncated:
        result += f"\n... (showing first {_MAX_RESULT_ROWS} rows)"
    if len(result) > _MAX_RESULT_CHARS:
        result = result[:_MAX_RESULT_CHARS] + "\n... (truncated)"
    return result


def _validate_read_only_query(query: str) -> str | None:
    """Return an error if query can access files or mutate DuckDB state."""
    stripped = query.strip()
    if not stripped:
        return _SQL_READ_ONLY_ERROR

    first_token = stripped.split(None, 1)[0].upper()
    if first_token not in {"SELECT", "WITH"}:
        return _SQL_READ_ONLY_ERROR

    for pattern in _BLOCKED_SQL_PATTERNS:
        if re.search(pattern, stripped, flags=re.IGNORECASE):
            return _SQL_READ_ONLY_ERROR
    return None


def _run_query(
    path: Path,
    query: str,
    output_path: Path | None,
    sheet_name: str | None = None,
) -> str:
    """Execute SQL query against a data file. Returns formatted result string."""
    import duckdb

    query_error = _validate_read_only_query(query)
    if query_error:
        return query_error

    conn = duckdb.connect(":memory:")
    try:
        ext = path.suffix.lower()
        # Escape single quotes in path to prevent SQL injection via file names.
        p = str(path).replace("'", "''")

        if ext == ".csv":
            enc = detect_csv_encoding(path)
            if enc in ("utf-8", "utf-8-sig"):
                conn.execute(
                    f"CREATE TABLE data AS SELECT * FROM read_csv_auto('{p}', encoding='utf-8')"
                )
            else:
                utf8_path = reencode_csv_to_utf8(path, enc)
                try:
                    up = str(utf8_path).replace("'", "''")
                    conn.execute(f"CREATE TABLE data AS SELECT * FROM read_csv_auto('{up}')")
                finally:
                    utf8_path.unlink(missing_ok=True)
        elif ext in (".json", ".jsonl", ".ndjson"):
            conn.execute(f"CREATE TABLE data AS SELECT * FROM read_json_auto('{p}')")
        elif ext == ".xlsx":
            if sheet_name == "all":
                # Load all sheets as separate tables
                all_sheets = load_all_xlsx_sheets(path)
                if not all_sheets:
                    return "Error: XLSX file is empty."
                first_name: str | None = None
                for sname, srows in all_sheets.items():
                    if not srows:
                        continue
                    if first_name is None:
                        first_name = sname
                    safe = _sanitize_table_name(sname)
                    _init_duckdb_table(conn, safe, srows)
                # Create 'data' alias for first sheet (backward compat)
                if first_name is not None:
                    safe_first = _sanitize_table_name(first_name)
                    conn.execute(f'CREATE VIEW data AS SELECT * FROM "{safe_first}"')
            elif sheet_name is not None:
                rows = load_xlsx_as_dicts(path, sheet_name=sheet_name)
                if not rows:
                    return f"Error: Sheet '{sheet_name}' is empty."
                init_duckdb_with_xlsx(path, rows, conn)
            else:
                rows = load_xlsx_as_dicts(path)
                if not rows:
                    return "Error: XLSX file is empty."
                init_duckdb_with_xlsx(path, rows, conn)
        elif ext == ".parquet":
            conn.execute(f"CREATE TABLE data AS SELECT * FROM read_parquet('{p}')")
        else:
            return f"Error: Unsupported file format '{ext}'. Use CSV, XLSX, JSON, or Parquet."

        # Get row count.
        count = conn.execute("SELECT COUNT(*) FROM data").fetchone()
        count_str = str(count[0]) if count else "0"

        # Execute user query.
        result = conn.execute(query)
        columns = [desc[0] for desc in result.description] if result.description else []
        rows = result.fetchall()

        output = _format_results(columns, rows)
        output = f"Source: {path.name} ({count_str} rows)\n\n{output}"

        # Save to CSV if requested.
        if output_path is not None:
            with open(output_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(columns)
                writer.writerows(rows)
            output += f"\n\nResults saved to {output_path}"

        return output
    except Exception as e:
        return f"Error: {e}"
    finally:
        conn.close()


class TableQueryTool(Tool):
    """Run SQL queries on tabular data files."""

    name = "table_query"
    description = (
        "Run SQL queries on tabular data files (CSV, XLSX, JSON, Parquet). "
        "The loaded data is available as table 'data'. "
        "Example: SELECT column, COUNT(*) FROM data GROUP BY column"
    )
    params = [
        ToolParam(
            name="path",
            type="string",
            description="Path to data file (CSV, XLSX, JSON, or Parquet)",
        ),
        ToolParam(
            name="query",
            type="string",
            description="SQL query to execute (table name: 'data')",
        ),
        ToolParam(
            name="sheet_name",
            type="string",
            description=(
                "Excel sheet name to load. Use 'all' to load all sheets as "
                "separate tables (named after sheets). Default: active sheet."
            ),
            required=False,
        ),
        ToolParam(
            name="output_path",
            type="string",
            description="Save query results to a CSV file",
            required=False,
        ),
    ]
    risk_level = RiskLevel.MEDIUM

    async def execute(self, **kwargs: Any) -> str:
        path_str = kwargs.get("path", "")
        query = kwargs.get("query", "")
        output_str = kwargs.get("output_path")
        sheet_name: str | None = kwargs.get("sheet_name")

        if not path_str:
            return "Error: 'path' is required."
        if not query:
            return "Error: 'query' is required."

        try:
            resolved = resolve_and_validate_path(path_str)
        except PermissionError as e:
            return f"Error: {e}"

        if not resolved.is_file():
            return f"Error: File not found: {path_str}"

        output_path: Path | None = None
        if output_str:
            try:
                output_path = resolve_and_validate_path(output_str)
            except PermissionError as e:
                return f"Error: {e}"

        return await run_in_thread(_run_query, resolved, query, output_path, sheet_name)
