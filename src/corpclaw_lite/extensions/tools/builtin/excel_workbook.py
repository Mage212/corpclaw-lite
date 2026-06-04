"""excel_workbook -- structured Excel read/fill operations for subagents.

Reads cells by coordinate and fills cells while preserving all formatting,
formulas, and merged ranges.  Designed for working with template-based
corporate reports where structure must be maintained.
"""

from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

from corpclaw_lite.extensions.tools.base import RiskLevel, Tool, ToolParam
from corpclaw_lite.extensions.tools.builtin.files import resolve_and_validate_path
from corpclaw_lite.utils.async_helpers import run_in_thread

__all__ = ["ExcelWorkbookTool"]

_MAX_DEFAULT_ROWS = 50
_MAX_ROWS_PER_CALL = 100
_MAX_OUTPUT_CHARS = 15_000
_FORMULA_MODES = {"both", "values", "formulas"}
_MISSING_CACHED_VALUE = "<unavailable>"


def _resolve_sheet(wb: Any, sheet_name: str | None) -> Any:
    """Get worksheet by name or fall back to active sheet."""
    if sheet_name is not None:
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        return wb[sheet_name]
    return wb.active


def _normalize_formula_mode(show_formulas: bool, formula_mode: Any) -> str:
    if formula_mode is None or formula_mode == "":
        return "formulas" if show_formulas else "both"
    mode = str(formula_mode).strip().lower()
    if mode not in _FORMULA_MODES:
        allowed = ", ".join(sorted(_FORMULA_MODES))
        raise ValueError(f"Invalid formula_mode '{formula_mode}'. Use one of: {allowed}.")
    return mode


def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _format_value(value: Any, max_chars: int, *, quote_strings: bool = False) -> str:
    if value is None:
        return "None"
    if isinstance(value, datetime):
        if value.time().replace(microsecond=0) == datetime.min.time():
            text = value.strftime("%d.%m.%Y")
        else:
            text = value.strftime("%d.%m.%Y %H:%M:%S")
    elif isinstance(value, date):
        text = value.strftime("%d.%m.%Y")
    elif isinstance(value, str) and quote_strings:
        text = repr(value)
    else:
        text = str(value)
    return text[:max_chars]


def _format_compact_cell(
    formula_cell: Any,
    value_cell: Any | None,
    formula_mode: str,
) -> str | None:
    formula_value = formula_cell.value
    value = value_cell.value if value_cell is not None else formula_value

    if formula_mode == "formulas":
        if formula_value is None:
            return None
        formula_hint = " (formula)" if _is_formula(formula_value) else ""
        return f"{formula_cell.coordinate}={_format_value(formula_value, 50)}{formula_hint}"

    if formula_mode == "values":
        if value is None:
            return None
        return f"{formula_cell.coordinate}={_format_value(value, 50)}"

    if _is_formula(formula_value):
        cached_value = _MISSING_CACHED_VALUE if value is None else _format_value(value, 50)
        return (
            f"{formula_cell.coordinate}=formula:{_format_value(formula_value, 50)} "
            f"| cached_value={cached_value}"
        )
    if value is None:
        return None
    return f"{formula_cell.coordinate}={_format_value(value, 50)}"


def _format_detail_cell(formula_cell: Any, value_cell: Any | None, formula_mode: str) -> str:
    formula_value = formula_cell.value
    value = value_cell.value if value_cell is not None else formula_value

    if formula_mode == "formulas":
        formula_hint = " (formula)" if _is_formula(formula_value) else ""
        return (
            f"  {formula_cell.coordinate} = "
            f"{_format_value(formula_value, 60, quote_strings=True)}{formula_hint}"
        )

    if formula_mode == "values":
        return f"  {formula_cell.coordinate} = {_format_value(value, 60, quote_strings=True)}"

    if _is_formula(formula_value):
        cached_value = (
            _MISSING_CACHED_VALUE if value is None else _format_value(value, 60, quote_strings=True)
        )
        return (
            f"  {formula_cell.coordinate} = formula:{_format_value(formula_value, 60)} "
            f"| cached_value={cached_value}"
        )
    return f"  {formula_cell.coordinate} = {_format_value(value, 60, quote_strings=True)}"


def _read_cells(
    path: Path,
    sheet_name: str | None,
    cells: str,
    formula_mode: str,
    offset: int,
    limit: int,
) -> str:
    import openpyxl

    wb_formula = openpyxl.load_workbook(str(path), data_only=False)
    wb_values = (
        openpyxl.load_workbook(str(path), data_only=True)
        if formula_mode in {"both", "values"}
        else None
    )
    try:
        ws = _resolve_sheet(wb_formula, sheet_name)
        ws_values = _resolve_sheet(wb_values, sheet_name) if wb_values is not None else None
        if ws is None:
            return "Error: Workbook has no sheets."

        lines: list[str] = [f'Sheet: "{ws.title}"']

        if not cells:
            # Default: rows with pagination, non-None only
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            start_row = offset + 1
            end_row = min(start_row + limit, max_row + 1)
            data_rows = 0

            for row_idx in range(start_row, end_row):
                row_data: list[str] = []
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value_cell = (
                        ws_values.cell(row=row_idx, column=col_idx)
                        if ws_values is not None
                        else None
                    )
                    formatted = _format_compact_cell(cell, value_cell, formula_mode)
                    if formatted is not None:
                        row_data.append(formatted)
                if row_data:
                    lines.append(f"  Row {row_idx}: {'  |  '.join(row_data)}")
                    data_rows += 1

            if data_rows == limit and end_row <= max_row:
                lines.append(
                    f"  Showing rows {start_row}-{end_row - 1}. "
                    f"More rows may exist — call again with offset={offset + limit}."
                )
            return "\n".join(lines)

        def append_range(range_ref: str) -> None:
            """Append a compact row-based view for one rectangular range."""
            lines.append(f"Range: {range_ref}")
            try:
                rows_in_range = list(ws[range_ref])
            except Exception as e:
                lines.append(f"  Error: invalid range '{range_ref}': {e}")
                return

            total_rows = len(rows_in_range)
            start_idx = min(offset, total_rows)
            end_idx = min(start_idx + limit, total_rows)
            data_rows = 0

            for row in rows_in_range[start_idx:end_idx]:
                row_data: list[str] = []
                for cell in row:
                    value_cell = ws_values[cell.coordinate] if ws_values is not None else None
                    formatted = _format_compact_cell(cell, value_cell, formula_mode)
                    if formatted is not None:
                        row_data.append(formatted)
                if row_data:
                    lines.append(f"  Row {row[0].row}: {'  |  '.join(row_data)}")
                    data_rows += 1

            if data_rows == limit and end_idx < total_rows:
                lines.append(
                    f"  Showing rows {start_idx + 1}-{end_idx} of range. "
                    f"More rows may exist — call again with offset={offset + limit}."
                )

        def append_cell(cell_ref: str) -> None:
            """Append a detailed view for one cell reference."""
            try:
                cell = ws[cell_ref]
            except Exception:
                lines.append(f"  {cell_ref} = Error: invalid cell reference")
                return
            if not hasattr(cell, "coordinate"):
                lines.append(f"  {cell_ref} = Error: invalid cell reference")
                return
            value_cell = ws_values[cell.coordinate] if ws_values is not None else None
            lines.append(_format_detail_cell(cell, value_cell, formula_mode))

        # Parse cells specification. Supports one range ("B2:F7"), one cell ("A1"), or
        # a comma-separated mix ("A1,B2:D4,F8:G9").
        for cell_spec in [part.strip() for part in cells.split(",") if part.strip()]:
            if ":" in cell_spec:
                append_range(cell_spec)
            else:
                append_cell(cell_spec)

        return "\n".join(lines)
    finally:
        wb_formula.close()
        if wb_values is not None:
            wb_values.close()


def _fill_cells(
    path: Path,
    sheet_name: str | None,
    cells_json: str | dict[str, Any],
    output_path: Path,
) -> str:
    import openpyxl
    from openpyxl.cell.cell import MergedCell

    if not cells_json:
        return 'Error: \'cells\' is required for fill action. Provide JSON like {"B2": "value"}.'

    # XML fallback parser may pre-deserialize JSON, delivering a dict directly.
    if isinstance(cells_json, dict):
        cells_dict: dict[str, Any] = cells_json
    else:
        try:
            cells_dict = json.loads(cells_json)
        except (json.JSONDecodeError, TypeError) as e:
            return f"Error: Invalid JSON in 'cells': {e}"

    if not cells_dict:
        return "Error: 'cells' must be a non-empty JSON object {\"address\": value}."

    # Load WITHOUT data_only and read_only to preserve formatting
    wb = openpyxl.load_workbook(str(path))
    try:
        ws = _resolve_sheet(wb, sheet_name)
        if ws is None:
            return "Error: Workbook has no sheets."

        filled: list[str] = []
        skipped_merged: list[str] = []
        for addr, value in cells_dict.items():
            try:
                cell = ws[addr]
                if isinstance(cell, MergedCell):
                    skipped_merged.append(addr)
                    continue
                cell.value = value
                filled.append(addr)
            except Exception as e:
                return f"Error writing to {addr}: {e}"

        try:
            wb.save(str(output_path))
        except Exception as e:
            return f"Error saving file: {e}"

        sheet_label = sheet_name or ws.title
        msg = (
            f"Filled {len(filled)} cells in sheet '{sheet_label}': {', '.join(filled)}. "
            f"Saved to {output_path.name}."
        )
        if skipped_merged:
            msg += (
                f" Skipped {len(skipped_merged)} merged cells"
                f" (non-top-left): {', '.join(skipped_merged)}."
            )
        return msg
    finally:
        wb.close()


class ExcelWorkbookTool(Tool):
    """Structured Excel operations: read cells by coordinate, fill cells
    preserving all formatting and formulas."""

    name = "excel_workbook"
    description = (
        "Read Excel cells by coordinate. Default: first 50 non-empty rows. "
        "Use 'cells' for specific ranges. Use 'offset'/'limit' for pagination "
        "(max 100 rows per call). If output shows 'More rows may exist', "
        "call again with increased offset to continue reading. For fill action, "
        "the default is a safe <name>_filled.xlsx copy; use in_place=true only "
        "when overwriting the original is explicitly requested."
    )
    params = [
        ToolParam(name="path", type="string", description="Path to .xlsx file"),
        ToolParam(
            name="action",
            type="string",
            description="Action: read or fill",
            enum=["read", "fill"],
        ),
        ToolParam(
            name="sheet_name",
            type="string",
            description="Sheet name (default: active/first sheet)",
            required=False,
        ),
        ToolParam(
            name="cells",
            type="string",
            description=(
                "Cell reference(s). For 'read': single cell ('A1'), range ('B2:F7'), "
                "or comma-separated mix ('A1,C3,D5:F8'). For 'fill': JSON object "
                '({"B2": "value", "G15": 150})'
            ),
            required=False,
        ),
        ToolParam(
            name="output_path",
            type="string",
            description=(
                "Output .xlsx path for fill action. Default: <name>_filled.xlsx. "
                "Ignored for read action."
            ),
            required=False,
        ),
        ToolParam(
            name="in_place",
            type="boolean",
            description="For fill action only: true to overwrite the original workbook explicitly",
            required=False,
        ),
        ToolParam(
            name="show_formulas",
            type="boolean",
            description=(
                "Backward-compatible read option. true means formula_mode='formulas'. "
                "Prefer formula_mode for new calls."
            ),
            required=False,
        ),
        ToolParam(
            name="formula_mode",
            type="string",
            description=(
                "Read formula handling: both (default: show formula cells with cached values), "
                "values (old value-only data_only read), or formulas (formula strings only). "
                "Cached values come from the workbook; formulas are not recalculated."
            ),
            enum=["both", "values", "formulas"],
            required=False,
        ),
        ToolParam(
            name="offset",
            type="integer",
            description="Row offset for pagination (0-based). Use to continue reading.",
            required=False,
        ),
        ToolParam(
            name="limit",
            type="integer",
            description="Max rows to return (default: 50, max: 100)",
            required=False,
        ),
    ]
    risk_level = RiskLevel.MEDIUM
    parallel_safe = False

    async def execute(self, **kwargs: Any) -> str:
        path_str = kwargs.get("path", "")
        action = kwargs.get("action", "")

        if not path_str:
            return "Error: 'path' is required."
        if not action:
            return "Error: 'action' is required (read or fill)."

        try:
            resolved = resolve_and_validate_path(path_str)
        except PermissionError as e:
            return f"Error: {e}"

        if not resolved.is_file():
            return f"Error: File not found: {path_str}"
        if resolved.suffix.lower() != ".xlsx":
            return "Error: Only .xlsx files are supported."

        sheet_name: str | None = kwargs.get("sheet_name")
        cells = kwargs.get("cells", "")
        show_formulas = kwargs.get("show_formulas", False)
        try:
            formula_mode = _normalize_formula_mode(show_formulas, kwargs.get("formula_mode"))
        except ValueError as e:
            return f"Error: {e}"
        offset = kwargs.get("offset", 0)
        limit = min(kwargs.get("limit", _MAX_DEFAULT_ROWS), _MAX_ROWS_PER_CALL)

        if action == "read":
            try:
                result = await run_in_thread(
                    _read_cells, resolved, sheet_name, cells, formula_mode, offset, limit
                )
                if len(result) > _MAX_OUTPUT_CHARS:
                    result = result[:_MAX_OUTPUT_CHARS] + "\n... (truncated)"
                return result
            except ValueError as e:
                return f"Error: {e}"
        if action == "fill":
            output_path = resolved.parent / f"{resolved.stem}_filled{resolved.suffix}"
            in_place = bool(kwargs.get("in_place", False))
            output_str = kwargs.get("output_path")
            if in_place:
                output_path = resolved
            elif output_str:
                try:
                    output_path = resolve_and_validate_path(output_str)
                except PermissionError as e:
                    return f"Error: {e}"
                if output_path.suffix.lower() != ".xlsx":
                    return "Error: output_path must end with .xlsx."
            try:
                return await run_in_thread(_fill_cells, resolved, sheet_name, cells, output_path)
            except ValueError as e:
                return f"Error: {e}"
        return f"Error: Unknown action '{action}'. Use 'read' or 'fill'."
