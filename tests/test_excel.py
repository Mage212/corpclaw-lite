"""Tests for NormalizeExcelTool — aligned with original normalize_format.py logic."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pytest

from corpclaw_lite.extensions.tools.builtin.excel import (
    NormalizeExcelTool,
    _clean_chars,
    _detect_column_type,
    _fix_value,
)


@pytest.fixture
def tool() -> NormalizeExcelTool:
    return NormalizeExcelTool()


def _create_xlsx(path: Path, headers: list[str], rows: list[list[object]]) -> Path:
    """Helper to create a .xlsx file with openpyxl."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    wb.save(str(path))
    return path


# --- Unit tests for helper functions ---


class TestCleanChars:
    def test_removes_zero_width_space(self) -> None:
        assert _clean_chars("hello\u200bworld") == "helloworld"

    def test_replaces_nbsp(self) -> None:
        assert _clean_chars("hello\u00a0world") == "hello world"

    def test_replaces_line_separator(self) -> None:
        assert _clean_chars("hello\u2028world") == "hello world"

    def test_strips_control_chars(self) -> None:
        assert _clean_chars("hello\x00\x01world") == "helloworld"

    def test_preserves_tab_and_newline(self) -> None:
        assert _clean_chars("hello\tworld\n") == "hello\tworld"


class TestDetectColumnType:
    def test_numeric_inventar(self) -> None:
        assert _detect_column_type("Зачетный инвентарь") == "numeric"

    def test_numeric_summa(self) -> None:
        assert _detect_column_type("Сумма с НДС") == "numeric"

    def test_inn(self) -> None:
        assert _detect_column_type("ИНН заказчика") == "inn"

    def test_date(self) -> None:
        assert _detect_column_type("Дата заключения договора") == "date"

    def test_text(self) -> None:
        assert _detect_column_type("Наименование") == "text"

    def test_case_insensitive(self) -> None:
        assert _detect_column_type("ИНН") == "inn"
        assert _detect_column_type("инн") == "inn"

    def test_numeric_priority_over_inn(self) -> None:
        assert _detect_column_type("Сумма") == "numeric"


class TestFixValueInn:
    def test_float_to_string(self) -> None:
        assert _fix_value(7700000000.0, "inn") == "7700000000"

    def test_int_to_string(self) -> None:
        assert _fix_value(7700000000, "inn") == "7700000000"

    def test_leading_zeros_10digit(self) -> None:
        result = _fix_value(123456789.0, "inn")
        assert result == "0123456789"

    def test_leading_zeros_12digit(self) -> None:
        result = _fix_value(12345678901.0, "inn")
        assert result == "012345678901"

    def test_scientific_notation_string(self) -> None:
        assert _fix_value("7.7E+09", "inn") == "7700000000"

    def test_regular_string(self) -> None:
        assert _fix_value("7700000000", "inn") == "7700000000"


class TestFixValueDate:
    def test_datetime_to_string(self) -> None:
        dt = datetime(2024, 3, 15)
        assert _fix_value(dt, "date") == "15.03.2024"

    def test_serial_date(self) -> None:
        result = _fix_value(45366.0, "date")
        assert result == "15.03.2024"

    def test_datetime_in_non_date_column(self) -> None:
        dt = datetime(2024, 3, 15)
        assert _fix_value(dt, "text") is None

    def test_string_passthrough(self) -> None:
        assert _fix_value("15.03.2024", "date") == "15.03.2024"


class TestFixValueNumeric:
    def test_float_rounding(self) -> None:
        assert _fix_value(19.247999999999998, "numeric") == 19.25

    def test_int_preserved(self) -> None:
        assert _fix_value(42, "numeric") == 42

    def test_string_with_dot(self) -> None:
        assert _fix_value("3.14", "numeric") == 3.14

    def test_string_without_dot(self) -> None:
        assert _fix_value("42", "numeric") == 42

    def test_string_with_comma(self) -> None:
        assert _fix_value("123,45", "numeric") == 123.45

    def test_nan_returns_none(self) -> None:
        assert _fix_value(float("nan"), "numeric") is None

    def test_inf_returns_none(self) -> None:
        assert _fix_value(float("inf"), "numeric") is None


class TestFixValueText:
    def test_float_whole_to_string(self) -> None:
        assert _fix_value(123.0, "text") == "123"

    def test_float_fractional_to_string(self) -> None:
        assert _fix_value(123.45, "text") == "123.45"

    def test_int_passthrough(self) -> None:
        assert _fix_value(42, "text") == "42"

    def test_none_passthrough(self) -> None:
        assert _fix_value(None, "text") is None

    def test_invisible_chars_removed(self) -> None:
        assert _fix_value("hello\u200bworld", "text") == "helloworld"

    def test_empty_string_to_none(self) -> None:
        assert _fix_value("   ", "text") is None


# --- Integration tests for the tool ---


@pytest.mark.asyncio
async def test_headers_preserved(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.chdir(tmp_path)
    _create_xlsx(
        tmp_path / "data.xlsx", ["Сумма с НДС", " ИНН заказчика "], [["1000", 7700000000.0]]
    )

    import openpyxl

    tool = NormalizeExcelTool()
    result = await tool.execute(path="data.xlsx")
    assert "Error" not in result

    wb = openpyxl.load_workbook(str(tmp_path / "data_normalized.xlsx"))
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "Сумма с НДС"
    assert ws.cell(row=1, column=2).value == "ИНН заказчика"


@pytest.mark.asyncio
async def test_remove_empty_rows(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.chdir(tmp_path)
    _create_xlsx(
        tmp_path / "data.xlsx",
        ["name", "age"],
        [["Alice", 30], [None, None], ["Bob", 25]],
    )

    tool = NormalizeExcelTool()
    result = await tool.execute(path="data.xlsx")
    assert "Empty rows removed: 1" in result

    import openpyxl

    wb = openpyxl.load_workbook(str(tmp_path / "data_normalized.xlsx"))
    ws = wb.active
    assert ws.max_row == 3


@pytest.mark.asyncio
async def test_duplicates_preserved(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.chdir(tmp_path)
    _create_xlsx(
        tmp_path / "data.xlsx",
        ["name", "age"],
        [["Alice", 30], ["Alice", 30], ["Bob", 25]],
    )

    tool = NormalizeExcelTool()
    result = await tool.execute(path="data.xlsx")
    assert "Duplicates" not in result

    import openpyxl

    wb = openpyxl.load_workbook(str(tmp_path / "data_normalized.xlsx"))
    ws = wb.active
    assert ws.max_row == 4


@pytest.mark.asyncio
async def test_output_formatting(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.chdir(tmp_path)
    _create_xlsx(tmp_path / "data.xlsx", ["Name", "ИНН"], [["Alice", 7700000000.0]])

    import openpyxl

    tool = NormalizeExcelTool()
    await tool.execute(path="data.xlsx")

    wb = openpyxl.load_workbook(str(tmp_path / "data_normalized.xlsx"))
    ws = wb.active

    assert ws.title == "Данные"
    assert ws.freeze_panes == "A2"

    header_cell = ws.cell(row=1, column=1)
    assert header_cell.font.bold is True
    assert header_cell.number_format == "@"

    inn_cell = ws.cell(row=2, column=2)
    assert inn_cell.number_format == "@"
    assert inn_cell.value == "7700000000"


@pytest.mark.asyncio
async def test_default_output_path(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.chdir(tmp_path)
    _create_xlsx(tmp_path / "report.xlsx", ["a"], [["x"]])

    tool = NormalizeExcelTool()
    result = await tool.execute(path="report.xlsx")
    assert "report_normalized.xlsx" in result
    assert (tmp_path / "report_normalized.xlsx").exists()


@pytest.mark.asyncio
async def test_custom_output_path(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.chdir(tmp_path)
    _create_xlsx(tmp_path / "data.xlsx", ["a"], [["x"]])

    tool = NormalizeExcelTool()
    result = await tool.execute(path="data.xlsx", output_path="clean.xlsx")
    assert "clean.xlsx" in result
    assert (tmp_path / "clean.xlsx").exists()


@pytest.mark.asyncio
async def test_file_not_found(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.chdir(tmp_path)
    tool = NormalizeExcelTool()
    result = await tool.execute(path="missing.xlsx")
    assert "Error" in result
    assert "does not exist" in result


@pytest.mark.asyncio
async def test_non_xlsx_error(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.chdir(tmp_path)
    (tmp_path / "data.csv").write_text("a,b\n1,2\n")
    tool = NormalizeExcelTool()
    result = await tool.execute(path="data.csv")
    assert "Error" in result
    assert ".xlsx" in result


@pytest.mark.asyncio
async def test_path_traversal_blocked(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.chdir(tmp_path)
    tool = NormalizeExcelTool()
    result = await tool.execute(path="../secret.xlsx")
    assert "Error" in result
