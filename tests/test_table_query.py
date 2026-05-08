"""Tests for TableQueryTool."""

from __future__ import annotations

import json
from pathlib import Path

import pytest

from corpclaw_lite.extensions.tools.builtin.table_query import TableQueryTool


@pytest.fixture
def tool() -> TableQueryTool:
    return TableQueryTool()


def _create_csv(path: Path, content: str) -> Path:
    path.write_text(content, encoding="utf-8")
    return path


def _create_xlsx(path: Path, headers: list[str], rows: list[list[object]]) -> Path:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, row in enumerate(rows, 2):
        for col, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col, value=val)
    wb.save(str(path))
    return path


def _create_json(path: Path, data: list[dict[str, object]]) -> Path:
    path.write_text(json.dumps(data), encoding="utf-8")
    return path


# --- Integration tests ---


class TestTableQueryTool:
    @pytest.mark.asyncio
    async def test_select_all_csv(
        self, tool: TableQueryTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        _create_csv(tmp_path / "data.csv", "name,age\nAlice,30\nBob,25\n")

        result = await tool.execute(path="data.csv", query="SELECT * FROM data")
        assert "Alice" in result
        assert "Bob" in result
        assert "2 rows" in result

    @pytest.mark.asyncio
    async def test_aggregation(
        self, tool: TableQueryTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        _create_csv(tmp_path / "data.csv", "dept,salary\nIT,100\nIT,200\nHR,150\n")

        result = await tool.execute(
            path="data.csv",
            query="SELECT dept, SUM(salary) as total FROM data GROUP BY dept ORDER BY dept",
        )
        assert "IT" in result
        assert "HR" in result

    @pytest.mark.asyncio
    async def test_where_clause(
        self, tool: TableQueryTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        _create_csv(tmp_path / "data.csv", "name,score\nAlice,95\nBob,60\nCarol,88\n")

        result = await tool.execute(
            path="data.csv",
            query="SELECT name FROM data WHERE score > 80",
        )
        assert "Alice" in result
        assert "Carol" in result
        assert "Bob" not in result

    @pytest.mark.asyncio
    async def test_xlsx_file(
        self, tool: TableQueryTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        _create_xlsx(
            tmp_path / "data.xlsx", ["city", "pop"], [["Moscow", 12000000], ["SPb", 5000000]]
        )

        result = await tool.execute(path="data.xlsx", query="SELECT * FROM data")
        assert "Moscow" in result
        assert "SPb" in result

    @pytest.mark.asyncio
    async def test_json_file(
        self, tool: TableQueryTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        _create_json(tmp_path / "data.json", [{"x": 1, "y": 2}, {"x": 3, "y": 4}])

        result = await tool.execute(path="data.json", query="SELECT SUM(x) as total_x FROM data")
        assert "4" in result

    @pytest.mark.asyncio
    async def test_output_path(
        self, tool: TableQueryTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        _create_csv(tmp_path / "data.csv", "name,val\nA,1\nB,2\n")

        result = await tool.execute(
            path="data.csv",
            query="SELECT * FROM data WHERE val > 0",
            output_path="result.csv",
        )
        assert "saved to" in result
        assert (tmp_path / "result.csv").exists()
        content = (tmp_path / "result.csv").read_text(encoding="utf-8")
        assert "A" in content

    @pytest.mark.asyncio
    async def test_sql_error(
        self, tool: TableQueryTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        _create_csv(tmp_path / "data.csv", "a,b\n1,2\n")

        result = await tool.execute(path="data.csv", query="INVALID SQL!!!")
        assert "Error" in result

    @pytest.mark.asyncio
    async def test_file_not_found(self, tool: TableQueryTool) -> None:
        result = await tool.execute(path="nonexistent.csv", query="SELECT 1")
        assert "Error" in result

    @pytest.mark.asyncio
    async def test_unsupported_format(
        self, tool: TableQueryTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        (tmp_path / "data.txt").write_text("hello", encoding="utf-8")

        result = await tool.execute(path="data.txt", query="SELECT 1")
        assert "Error" in result
        assert "Unsupported" in result

    @pytest.mark.asyncio
    async def test_empty_result(
        self, tool: TableQueryTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        _create_csv(tmp_path / "data.csv", "name,val\nAlice,1\nBob,2\n")

        result = await tool.execute(
            path="data.csv",
            query="SELECT * FROM data WHERE val > 100",
        )
        assert "0 rows" in result

    @pytest.mark.asyncio
    async def test_missing_params(self, tool: TableQueryTool) -> None:
        result = await tool.execute()
        assert "Error" in result

    # --- Cyrillic / encoding tests ---

    @pytest.mark.asyncio
    async def test_cp1251_csv_query(
        self, tool: TableQueryTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Windows-1251 CSV must be readable by DuckDB."""
        monkeypatch.chdir(tmp_path)
        content = "имя,зарплата\nАлексей,100000\nМария,120000\n"
        (tmp_path / "salaries.csv").write_bytes(content.encode("cp1251"))

        result = await tool.execute(path="salaries.csv", query="SELECT * FROM data")
        assert "Алексей" in result
        assert "Мария" in result

    @pytest.mark.asyncio
    async def test_cp1251_csv_aggregation(
        self, tool: TableQueryTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Aggregation on Windows-1251 CSV with Cyrillic column names."""
        monkeypatch.chdir(tmp_path)
        content = (
            "отдел,сотрудник,оклад\n"
            "Бухгалтерия,Иванов,80000\n"
            "IT,Сидоров,120000\n"
            "Бухгалтерия,Петрова,90000\n"
        )
        (tmp_path / "depts.csv").write_bytes(content.encode("cp1251"))

        result = await tool.execute(
            path="depts.csv",
            query="SELECT отдел, SUM(оклад) as итого FROM data GROUP BY отдел ORDER BY отдел",
        )
        assert "Бухгалтерия" in result
        assert "170000" in result  # 80000 + 90000

    @pytest.mark.asyncio
    async def test_utf8_bom_csv_query(
        self, tool: TableQueryTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """UTF-8 BOM CSV must be readable by DuckDB."""
        monkeypatch.chdir(tmp_path)
        content = "товар,количество\nМолоко,10\nХлеб,25\n"
        (tmp_path / "goods.csv").write_bytes(b"\xef\xbb\xbf" + content.encode("utf-8"))

        result = await tool.execute(path="goods.csv", query="SELECT * FROM data")
        assert "Молоко" in result
        assert "Хлеб" in result

    # --- sheet_name tests ---

    @pytest.mark.asyncio
    async def test_xlsx_specific_sheet(
        self, tool: TableQueryTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Query a specific named sheet in a multi-sheet workbook."""
        monkeypatch.chdir(tmp_path)
        import openpyxl

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sales"
        ws1.append(["product", "amount"])
        ws1.append(["Widget", 100])
        ws2 = wb.create_sheet("Returns")
        ws2.append(["product", "qty"])
        ws2.append(["Widget", 5])
        wb.save(str(tmp_path / "multi.xlsx"))

        result = await tool.execute(
            path="multi.xlsx",
            query="SELECT * FROM data",
            sheet_name="Returns",
        )
        assert "5" in result
        assert "100" not in result

    @pytest.mark.asyncio
    async def test_xlsx_all_sheets(
        self, tool: TableQueryTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """sheet_name='all' loads all sheets as separate tables."""
        monkeypatch.chdir(tmp_path)
        import openpyxl

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Q1"
        ws1.append(["city", "revenue"])
        ws1.append(["Moscow", 500])
        ws2 = wb.create_sheet("Q2")
        ws2.append(["city", "revenue"])
        ws2.append(["Moscow", 700])
        wb.save(str(tmp_path / "quarters.xlsx"))

        result = await tool.execute(
            path="quarters.xlsx",
            query="SELECT * FROM q2",
            sheet_name="all",
        )
        assert "700" in result

    @pytest.mark.asyncio
    async def test_xlsx_all_sheets_data_alias(
        self, tool: TableQueryTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """With sheet_name='all', 'data' VIEW aliases the first sheet."""
        monkeypatch.chdir(tmp_path)
        import openpyxl

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Alpha"
        ws1.append(["x"])
        ws1.append([42])
        ws2 = wb.create_sheet("Beta")
        ws2.append(["x"])
        ws2.append([99])
        wb.save(str(tmp_path / "two.xlsx"))

        result = await tool.execute(
            path="two.xlsx",
            query="SELECT * FROM data",
            sheet_name="all",
        )
        assert "42" in result

    @pytest.mark.asyncio
    async def test_xlsx_invalid_sheet_name(
        self, tool: TableQueryTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Invalid sheet name returns an error."""
        monkeypatch.chdir(tmp_path)
        _create_xlsx(tmp_path / "data.xlsx", ["a"], [[1]])

        result = await tool.execute(
            path="data.xlsx",
            query="SELECT * FROM data",
            sheet_name="NoSuchSheet",
        )
        assert "Error" in result
        assert "NoSuchSheet" in result
