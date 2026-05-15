"""Tests for ExcelWorkbookTool."""

from __future__ import annotations

import json
import zipfile
from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import PatternFill

from corpclaw_lite.extensions.tools.builtin.excel_workbook import ExcelWorkbookTool


@pytest.fixture
def tool() -> ExcelWorkbookTool:
    return ExcelWorkbookTool()


def _create_basic_xlsx(
    path: Path,
    headers: list[str],
    rows: list[list[object]],
) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, row in enumerate(rows, 2):
        for col, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col, value=val)
    wb.save(str(path))
    return path


def _inject_cached_formula_value(path: Path, cell: str, formula: str, cached_value: str) -> None:
    """Patch worksheet XML to simulate a workbook last recalculated by Excel/LibreOffice."""
    patched = path.with_suffix(".patched.xlsx")
    old_xml = f'<c r="{cell}"><f>{formula}</f><v /></c>'
    new_xml = f'<c r="{cell}"><f>{formula}</f><v>{cached_value}</v></c>'

    with zipfile.ZipFile(path) as zin, zipfile.ZipFile(patched, "w") as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                text = data.decode("utf-8")
                if old_xml not in text:
                    raise AssertionError(f"Could not find formula XML for {cell}")
                data = text.replace(old_xml, new_xml).encode("utf-8")
            zout.writestr(item, data)
    patched.replace(path)


class TestExcelWorkbookRead:
    @pytest.mark.asyncio
    async def test_read_default(
        self, tool: ExcelWorkbookTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Read without specifying cells returns first 50 rows."""
        monkeypatch.chdir(tmp_path)
        headers = ["Name", "Age", "City"]
        rows = [[f"user{i}", 20 + i, f"city{i}"] for i in range(1, 55)]
        _create_basic_xlsx(tmp_path / "data.xlsx", headers, rows)

        result = await tool.execute(path="data.xlsx", action="read")
        assert "Sheet:" in result
        # Default limit is 50, so rows 1-50 visible
        assert "Row 1:" in result
        assert "Row 50:" in result
        assert "Row 52:" not in result
        assert "Name" in result
        assert "user1" in result

    @pytest.mark.asyncio
    async def test_read_single_cell(
        self, tool: ExcelWorkbookTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Read a specific cell by coordinate."""
        monkeypatch.chdir(tmp_path)
        _create_basic_xlsx(tmp_path / "data.xlsx", ["Header1", "Header2"], [["val_a", 42]])

        result = await tool.execute(path="data.xlsx", action="read", cells="B2")
        assert "B2" in result
        assert "42" in result

    @pytest.mark.asyncio
    async def test_read_range(
        self, tool: ExcelWorkbookTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Read a range of cells like A1:C3 — compact format, no None."""
        monkeypatch.chdir(tmp_path)
        _create_basic_xlsx(
            tmp_path / "data.xlsx",
            ["H1", "H2", "H3"],
            [["a", "b", "c"], ["d", "e", "f"]],
        )

        result = await tool.execute(path="data.xlsx", action="read", cells="A1:C3")
        assert "Range: A1:C3" in result
        # Compact row-based format
        assert "Row 1:" in result
        assert "A1=H1" in result
        assert "C3=f" in result

    @pytest.mark.asyncio
    async def test_read_comma_separated(
        self, tool: ExcelWorkbookTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Read comma-separated individual cells."""
        monkeypatch.chdir(tmp_path)
        _create_basic_xlsx(
            tmp_path / "data.xlsx",
            ["Col1", "Col2", "Col3"],
            [["v1", "v2", "v3"]],
        )

        result = await tool.execute(path="data.xlsx", action="read", cells="A1,B2,C3")
        assert "A1" in result
        assert "B2" in result
        assert "C3" in result
        assert "'Col1'" in result
        assert "'v2'" in result

    @pytest.mark.asyncio
    async def test_read_specific_sheet(
        self, tool: ExcelWorkbookTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Read from a named sheet."""
        monkeypatch.chdir(tmp_path)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "DefaultSheet"
        ws2 = wb.create_sheet("DataSheet")
        ws2["A1"] = "OnDataSheet"
        ws2["B2"] = 99
        wb.save(str(tmp_path / "multi.xlsx"))

        result = await tool.execute(
            path="multi.xlsx", action="read", sheet_name="DataSheet", cells="A1"
        )
        assert "DataSheet" in result
        assert "'OnDataSheet'" in result

    @pytest.mark.asyncio
    async def test_read_show_formulas(
        self, tool: ExcelWorkbookTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Read with show_formulas=True shows formula strings."""
        monkeypatch.chdir(tmp_path)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = 10
        ws["B1"] = 20
        ws["C1"] = "=A1+B1"
        wb.save(str(tmp_path / "formula.xlsx"))

        # Default read shows formula plus cached value status.
        result_values = await tool.execute(path="formula.xlsx", action="read", cells="C1")
        assert "C1" in result_values
        assert "formula:=A1+B1" in result_values
        assert "cached_value=<unavailable>" in result_values

        # show_formulas=True — formula string shown
        result_formulas = await tool.execute(
            path="formula.xlsx", action="read", cells="C1", show_formulas=True
        )
        assert "C1" in result_formulas
        assert "=A1+B1" in result_formulas
        assert "(formula)" in result_formulas

    @pytest.mark.asyncio
    async def test_read_formula_mode_both_shows_cached_value(
        self, tool: ExcelWorkbookTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Default formula-aware mode shows formulas together with cached workbook values."""
        monkeypatch.chdir(tmp_path)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = 10
        ws["B1"] = 20
        ws["C1"] = "=A1+B1"
        wb.save(str(tmp_path / "cached_formula.xlsx"))
        wb.close()
        _inject_cached_formula_value(tmp_path / "cached_formula.xlsx", "C1", "A1+B1", "30")

        result = await tool.execute(path="cached_formula.xlsx", action="read", cells="C1")

        assert "C1" in result
        assert "formula:=A1+B1" in result
        assert "cached_value=30" in result

    @pytest.mark.asyncio
    async def test_read_formula_mode_values_keeps_old_value_only_behavior(
        self, tool: ExcelWorkbookTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """formula_mode=values keeps the old data_only read behavior for callers that need it."""
        monkeypatch.chdir(tmp_path)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = 10
        ws["B1"] = 20
        ws["C1"] = "=A1+B1"
        wb.save(str(tmp_path / "cached_formula.xlsx"))
        wb.close()
        _inject_cached_formula_value(tmp_path / "cached_formula.xlsx", "C1", "A1+B1", "30")

        result = await tool.execute(
            path="cached_formula.xlsx", action="read", cells="C1", formula_mode="values"
        )

        assert "C1 = 30" in result
        assert "formula:" not in result
        assert "cached_value=" not in result

    @pytest.mark.asyncio
    async def test_read_range_includes_formula_without_cached_value(
        self, tool: ExcelWorkbookTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Range mode must not hide formula cells just because cached value is missing."""
        monkeypatch.chdir(tmp_path)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Date"
        ws["A2"] = "=A1+1"
        wb.save(str(tmp_path / "missing_cached.xlsx"))
        wb.close()

        result = await tool.execute(path="missing_cached.xlsx", action="read", cells="A1:A2")

        assert "Row 2:" in result
        assert "A2=formula:=A1+1" in result
        assert "cached_value=<unavailable>" in result

    @pytest.mark.asyncio
    async def test_read_range_skips_none(
        self, tool: ExcelWorkbookTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Range mode skips None cells — compact row-based output."""
        monkeypatch.chdir(tmp_path)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["E1"] = "Label"
        ws["F1"] = "Value"
        ws["E2"] = "Brand"
        ws["F2"] = "TestCorp"
        wb.save(str(tmp_path / "sparse.xlsx"))

        result = await tool.execute(path="sparse.xlsx", action="read", cells="A1:F2")
        assert "Range: A1:F2" in result
        # None cells in columns A-D should not appear
        assert "A1=" not in result
        assert "B1=" not in result
        # Non-None cells should appear
        assert "E1=Label" in result
        assert "F1=Value" in result
        assert "E2=Brand" in result
        assert "F2=TestCorp" in result

    @pytest.mark.asyncio
    async def test_read_multiple_ranges(
        self, tool: ExcelWorkbookTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Read comma-separated ranges in one call."""
        monkeypatch.chdir(tmp_path)
        _create_basic_xlsx(
            tmp_path / "multi_range.xlsx",
            ["H1", "H2", "H3"],
            [["a", "b", "c"], ["d", "e", "f"]],
        )

        result = await tool.execute(path="multi_range.xlsx", action="read", cells="A1:B2,C2:C3")

        assert "Range: A1:B2" in result
        assert "Range: C2:C3" in result
        assert "A1=H1" in result
        assert "B2=b" in result
        assert "C2=c" in result
        assert "C3=f" in result

    @pytest.mark.asyncio
    async def test_read_mixed_cells_and_ranges(
        self, tool: ExcelWorkbookTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Read a comma-separated mix of single cells and ranges."""
        monkeypatch.chdir(tmp_path)
        _create_basic_xlsx(
            tmp_path / "mixed.xlsx",
            ["H1", "H2", "H3"],
            [["a", "b", "c"], ["d", "e", "f"]],
        )

        result = await tool.execute(path="mixed.xlsx", action="read", cells="A1,B2:C3")

        assert "A1 = 'H1'" in result
        assert "Range: B2:C3" in result
        assert "B2=b" in result
        assert "C3=f" in result

    @pytest.mark.asyncio
    async def test_read_invalid_range_in_list_returns_error(
        self, tool: ExcelWorkbookTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Invalid comma-list entries return readable errors instead of raising."""
        monkeypatch.chdir(tmp_path)
        _create_basic_xlsx(tmp_path / "invalid_range.xlsx", ["H1"], [["a"]])

        result = await tool.execute(
            path="invalid_range.xlsx", action="read", cells="A1,not-a-range:B2"
        )

        assert "A1 = 'H1'" in result
        assert "Error: invalid range 'not-a-range:B2'" in result

    @pytest.mark.asyncio
    async def test_pagination_offset_and_limit(
        self, tool: ExcelWorkbookTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Pagination: offset and limit control which rows are returned."""
        monkeypatch.chdir(tmp_path)
        headers = ["ID"]
        rows = [[i] for i in range(1, 11)]
        _create_basic_xlsx(tmp_path / "paged.xlsx", headers, rows)

        # Read first 3 rows (rows 1-3, offset=0, limit=3)
        result = await tool.execute(path="paged.xlsx", action="read", offset=0, limit=3)
        assert "Row 1:" in result
        assert "Row 3:" in result
        assert "Row 4:" not in result
        # Should show continuation hint
        assert "More rows may exist" in result

        # Read next page (rows 4-6, offset=3, limit=3)
        result2 = await tool.execute(path="paged.xlsx", action="read", offset=3, limit=3)
        assert "Row 4:" in result2
        assert "Row 6:" in result2
        assert "Row 3:" not in result2

        # Read last page (rows 8+, offset=7, limit=3) — 3 rows left + row 11 beyond
        result3 = await tool.execute(path="paged.xlsx", action="read", offset=7, limit=3)
        assert "Row 8:" in result3
        assert "Row 10:" in result3
        assert "Row 11:" not in result3
        # Still shows hint because limit reached
        assert "More rows may exist" in result3

        # Read final page (offset=10, limit=3) — only row 11 left
        result4 = await tool.execute(path="paged.xlsx", action="read", offset=10, limit=3)
        assert "Row 11:" in result4
        # Less than limit rows returned — no continuation hint
        assert "More rows may exist" not in result4

    @pytest.mark.asyncio
    async def test_pagination_range_mode(
        self, tool: ExcelWorkbookTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Pagination works in range mode too."""
        monkeypatch.chdir(tmp_path)
        headers = ["Val"]
        rows = [[i] for i in range(1, 21)]
        _create_basic_xlsx(tmp_path / "ranged.xlsx", headers, rows)

        result = await tool.execute(
            path="ranged.xlsx", action="read", cells="A1:A21", offset=0, limit=5
        )
        assert "Row 1:" in result
        assert "Row 5:" in result
        assert "Row 6:" not in result
        assert "More rows may exist" in result

    @pytest.mark.asyncio
    async def test_max_limit_capped_at_100(
        self, tool: ExcelWorkbookTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Limit > 100 is capped to 100."""
        monkeypatch.chdir(tmp_path)
        headers = ["Val"]
        rows = [[i] for i in range(1, 201)]
        _create_basic_xlsx(tmp_path / "big.xlsx", headers, rows)

        result = await tool.execute(path="big.xlsx", action="read", limit=999)
        assert "Row 100:" in result
        assert "Row 101:" not in result
        assert "More rows may exist" in result


class TestExcelWorkbookFill:
    @pytest.mark.asyncio
    async def test_fill_cells(
        self, tool: ExcelWorkbookTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Fill cells with JSON values and verify they are saved."""
        monkeypatch.chdir(tmp_path)
        _create_basic_xlsx(tmp_path / "fill.xlsx", ["H1", "H2"], [["old1", "old2"]])

        cells_json = json.dumps({"A2": "new_val", "B2": 100})
        result = await tool.execute(path="fill.xlsx", action="fill", cells=cells_json)
        assert "Filled 2 cells" in result
        assert "Saved to fill_filled.xlsx" in result

        # Verify values persisted
        wb = openpyxl.load_workbook(str(tmp_path / "fill_filled.xlsx"))
        ws = wb.active
        assert ws["A2"].value == "new_val"
        assert ws["B2"].value == 100
        wb.close()
        original = openpyxl.load_workbook(str(tmp_path / "fill.xlsx"))
        assert original.active["A2"].value == "old1"
        original.close()

    @pytest.mark.asyncio
    async def test_fill_preserves_formatting(
        self, tool: ExcelWorkbookTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Fill does not erase existing cell formatting."""
        monkeypatch.chdir(tmp_path)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Colored"
        ws["A1"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        ws["B1"] = "Other"
        wb.save(str(tmp_path / "fmt.xlsx"))

        result = await tool.execute(
            path="fmt.xlsx", action="fill", cells=json.dumps({"B1": "updated"})
        )
        assert "Filled 1 cells" in result

        # Verify A1's fill color is preserved
        wb2 = openpyxl.load_workbook(str(tmp_path / "fmt_filled.xlsx"))
        ws2 = wb2.active
        assert ws2["A1"].fill.start_color.rgb == "00FF0000"
        assert ws2["B1"].value == "updated"
        wb2.close()

    @pytest.mark.asyncio
    async def test_fill_preserves_merged(
        self, tool: ExcelWorkbookTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Fill does not destroy merged cell ranges."""
        monkeypatch.chdir(tmp_path)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.merge_cells("A1:B1")
        ws["A1"] = "MergedHeader"
        ws["C1"] = "Plain"
        wb.save(str(tmp_path / "merged.xlsx"))

        result = await tool.execute(
            path="merged.xlsx", action="fill", cells=json.dumps({"C1": "filled"})
        )
        assert "Filled 1 cells" in result

        # Verify merged range is preserved
        wb2 = openpyxl.load_workbook(str(tmp_path / "merged_filled.xlsx"))
        ws2 = wb2.active
        assert "A1:B1" in [str(m) for m in ws2.merged_cells.ranges]
        assert ws2["C1"].value == "filled"
        wb2.close()

    @pytest.mark.asyncio
    async def test_fill_invalid_json(
        self, tool: ExcelWorkbookTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Malformed JSON in cells parameter returns an error."""
        monkeypatch.chdir(tmp_path)
        _create_basic_xlsx(tmp_path / "bad.xlsx", ["A"], [[1]])

        result = await tool.execute(path="bad.xlsx", action="fill", cells="{not valid json}")
        assert "Error" in result
        assert "Invalid JSON" in result

    @pytest.mark.asyncio
    async def test_fill_accepts_dict_cells(
        self, tool: ExcelWorkbookTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Fill accepts a Python dict directly (XML fallback path)."""
        monkeypatch.chdir(tmp_path)
        _create_basic_xlsx(tmp_path / "dict.xlsx", ["H1", "H2"], [["old1", "old2"]])

        result = await tool.execute(
            path="dict.xlsx", action="fill", cells={"A2": "from_dict", "B2": 42}
        )
        assert "Filled 2 cells" in result
        assert "Saved to dict_filled.xlsx" in result

        wb = openpyxl.load_workbook(str(tmp_path / "dict_filled.xlsx"))
        ws = wb.active
        assert ws["A2"].value == "from_dict"
        assert ws["B2"].value == 42
        wb.close()

    @pytest.mark.asyncio
    async def test_fill_skips_merged_cells(
        self, tool: ExcelWorkbookTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Fill skips non-top-left merged cells and reports them."""
        monkeypatch.chdir(tmp_path)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.merge_cells("A1:C1")
        ws["A1"] = "MergedHeader"
        ws["D1"] = "Plain"
        wb.save(str(tmp_path / "merged_fill.xlsx"))

        result = await tool.execute(
            path="merged_fill.xlsx",
            action="fill",
            cells=json.dumps({"A1": "top_ok", "B1": "skip_me", "D1": "plain_ok"}),
        )
        assert "Filled 2 cells" in result
        assert "Skipped 1 merged cells" in result
        assert "B1" in result

        wb2 = openpyxl.load_workbook(str(tmp_path / "merged_fill_filled.xlsx"))
        ws2 = wb2.active
        assert ws2["A1"].value == "top_ok"
        assert ws2["D1"].value == "plain_ok"
        assert "A1:C1" in [str(m) for m in ws2.merged_cells.ranges]
        wb2.close()

    @pytest.mark.asyncio
    async def test_fill_all_merged_returns_skip_message(
        self, tool: ExcelWorkbookTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """When all target cells are merged non-top-left, report skip with no fills."""
        monkeypatch.chdir(tmp_path)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.merge_cells("A1:B1")
        ws["A1"] = "Header"
        wb.save(str(tmp_path / "all_merged.xlsx"))

        result = await tool.execute(
            path="all_merged.xlsx",
            action="fill",
            cells=json.dumps({"B1": "should_skip"}),
        )
        assert "Filled 0 cells" in result
        assert "Skipped 1 merged cells" in result

    @pytest.mark.asyncio
    async def test_fill_in_place_explicit(
        self, tool: ExcelWorkbookTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """in_place=true overwrites the original workbook explicitly."""
        monkeypatch.chdir(tmp_path)
        _create_basic_xlsx(tmp_path / "inplace.xlsx", ["H1"], [["old"]])

        result = await tool.execute(
            path="inplace.xlsx",
            action="fill",
            cells=json.dumps({"A2": "new"}),
            in_place=True,
        )

        assert "Saved to inplace.xlsx" in result
        wb = openpyxl.load_workbook(str(tmp_path / "inplace.xlsx"))
        assert wb.active["A2"].value == "new"
        wb.close()

    @pytest.mark.asyncio
    async def test_fill_custom_output_path(
        self, tool: ExcelWorkbookTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Fill can write to a caller-provided output workbook."""
        monkeypatch.chdir(tmp_path)
        _create_basic_xlsx(tmp_path / "template.xlsx", ["H1"], [["old"]])

        result = await tool.execute(
            path="template.xlsx",
            action="fill",
            cells=json.dumps({"A2": "new"}),
            output_path="result.xlsx",
        )

        assert "Saved to result.xlsx" in result
        assert (tmp_path / "result.xlsx").exists()
        original = openpyxl.load_workbook(str(tmp_path / "template.xlsx"))
        assert original.active["A2"].value == "old"
        original.close()


class TestExcelWorkbookErrors:
    @pytest.mark.asyncio
    async def test_path_traversal_blocked(
        self, tool: ExcelWorkbookTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Path traversal with ../ is blocked."""
        monkeypatch.chdir(tmp_path)
        result = await tool.execute(path="../../../etc/passwd", action="read")
        assert "Error" in result

    @pytest.mark.asyncio
    async def test_non_xlsx_error(
        self, tool: ExcelWorkbookTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Non-.xlsx file extension returns an error."""
        monkeypatch.chdir(tmp_path)
        (tmp_path / "data.csv").write_text("a,b\n1,2", encoding="utf-8")

        result = await tool.execute(path="data.csv", action="read")
        assert "Error" in result
        assert ".xlsx" in result

    @pytest.mark.asyncio
    async def test_fill_output_path_outside_workspace_blocked(
        self, tool: ExcelWorkbookTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Fill output_path cannot escape the workspace."""
        workspace = tmp_path / "workspace"
        workspace.mkdir()
        monkeypatch.chdir(workspace)
        _create_basic_xlsx(workspace / "data.xlsx", ["A"], [[1]])

        result = await tool.execute(
            path="data.xlsx",
            action="fill",
            cells=json.dumps({"A2": "new"}),
            output_path="../escaped.xlsx",
        )

        assert "Error" in result
        assert not (tmp_path / "escaped.xlsx").exists()

    @pytest.mark.asyncio
    async def test_unknown_action_error(
        self, tool: ExcelWorkbookTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Invalid action returns an error."""
        monkeypatch.chdir(tmp_path)
        _create_basic_xlsx(tmp_path / "data.xlsx", ["A"], [[1]])

        result = await tool.execute(path="data.xlsx", action="delete")
        assert "Error" in result
        assert "Unknown action" in result

    @pytest.mark.asyncio
    async def test_missing_params(self, tool: ExcelWorkbookTool) -> None:
        """Missing required params returns errors."""
        result_no_path = await tool.execute(action="read")
        assert "Error" in result_no_path
        assert "path" in result_no_path

        result_no_action = await tool.execute(path="file.xlsx")
        assert "Error" in result_no_action
        assert "action" in result_no_action

    @pytest.mark.asyncio
    async def test_invalid_sheet_name(
        self, tool: ExcelWorkbookTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Reading from a nonexistent sheet returns error string."""
        monkeypatch.chdir(tmp_path)
        _create_basic_xlsx(tmp_path / "data.xlsx", ["A"], [[1]])

        result = await tool.execute(path="data.xlsx", action="read", sheet_name="NoSuchSheet")
        assert "Error" in result
        assert "NoSuchSheet" in result
        assert "not found" in result
