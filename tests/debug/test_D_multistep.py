"""Group D — Multi-step workflow integration tests.

Each test verifies a realistic work scenario that requires the agent to
chain multiple tool calls across ≥ 3 ReAct iterations.

Run:
    uv run pytest tests/debug/test_D_multistep.py -v -s
"""

from __future__ import annotations

from pathlib import Path

import pytest

from corpclaw_lite.agent.loop import AgentLoop
from corpclaw_lite.extensions.tools.registry import ToolRegistry
from corpclaw_lite.users.models import User

from .helpers import DebugAssertions, summarise_run

pytestmark = [pytest.mark.integration, pytest.mark.llm_required]


# ---------------------------------------------------------------------------
# D1 — Create → Edit → Read chain (3 tools, 3+ iterations)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_D1_create_edit_read_chain(
    agent_stack_no_container: tuple[AgentLoop, ToolRegistry],
    test_user: User,
    tmp_workspace: Path,
) -> None:
    """Agent creates a file, edits it, then reads it back — 3 sequential tools."""
    loop, _ = agent_stack_no_container

    reply, stats = await loop.run(
        test_user,
        "Выполни по порядку:\n"
        "1. Создай файл config.txt с содержимым: version=1\n"
        "2. Замени в нём 'version=1' на 'version=2'\n"
        "3. Прочитай файл и скажи его содержимое",
    )

    DebugAssertions.assert_tool_used(stats, "write_file", "edit_file", "read_file")
    DebugAssertions.assert_status_ok(stats)
    DebugAssertions.assert_min_iterations(stats, 3)
    DebugAssertions.assert_file_exists(tmp_workspace / "config.txt", contains="version=2")
    print(f"\n[D1] {summarise_run(reply, stats)}")


# ---------------------------------------------------------------------------
# D2 — Create directory structure + report file
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_D2_report_generation(
    agent_stack_no_container: tuple[AgentLoop, ToolRegistry],
    test_user: User,
    tmp_workspace: Path,
) -> None:
    """Agent creates a structured report file, then reads it back."""
    loop, _ = agent_stack_no_container

    reply, stats = await loop.run(
        test_user,
        "Создай файл reports/weekly.md со следующим содержимым:\n"
        "# Недельный отчёт\n\n"
        "1. Задача выполнена\n"
        "2. Ошибок нет\n"
        "3. Следующий шаг: деплой\n\n"
        "Затем прочитай файл и подтверди, что он создан.",
    )

    DebugAssertions.assert_tool_used(stats, "write_file")
    DebugAssertions.assert_status_ok(stats)

    report = tmp_workspace / "reports" / "weekly.md"
    DebugAssertions.assert_file_exists(report)
    content = report.read_text(encoding="utf-8")
    assert "Недельный отчёт" in content or "отчёт" in content.lower(), (
        "Report file missing expected heading"
    )
    print(f"\n[D2] {summarise_run(reply, stats)}")


# ---------------------------------------------------------------------------
# D3 — Search across multiple files, read matches, summarise
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_D3_search_and_summarise(
    agent_stack_no_container: tuple[AgentLoop, ToolRegistry],
    test_user: User,
    tmp_workspace: Path,
) -> None:
    """Agent searches for a token, reads matching files, writes a summary."""
    loop, _ = agent_stack_no_container

    # Pre-create files
    for i in range(1, 4):
        (tmp_workspace / f"report_{i}.txt").write_text(
            f"IMPORTANT_D03\nReport #{i}\nContent here.", encoding="utf-8"
        )
    (tmp_workspace / "other.txt").write_text("Nothing important.", encoding="utf-8")

    reply, stats = await loop.run(
        test_user,
        "Найди все файлы содержащие слово IMPORTANT_D03, прочитай их "
        "и напиши краткую сводку: сколько файлов найдено и что в них.",
    )

    DebugAssertions.assert_tool_used(stats, "search_files")
    DebugAssertions.assert_status_ok(stats)
    DebugAssertions.assert_min_iterations(stats, 2)
    # At least one report name must appear.
    # Normalize markdown-escaped underscores (report\_1 → report_1) before check.
    reply_normalized = reply.replace("\\_ ", "_").replace("\\_ ", "_").replace("\\_", "_")
    assert any(f"report_{i}" in reply_normalized for i in range(1, 4)), (
        f"Reply should mention found report files.\nReply: {reply[:400]}"
    )
    print(f"\n[D3] {summarise_run(reply, stats)}")


# ---------------------------------------------------------------------------
# D4 — Memory persists across separate loop.run() calls
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_D4_memory_across_runs(
    agent_stack_no_container: tuple[AgentLoop, ToolRegistry],
    test_user: User,
) -> None:
    """Fact stored in first run() is recalled in second run() for same user."""
    loop, _ = agent_stack_no_container

    unique_token = "SESSION_TOKEN_D04_XYZ"

    # First turn: store
    _, stats1 = await loop.run(
        test_user,
        f"Запомни этот токен: {unique_token}. Используй инструмент memory_store.",
    )
    DebugAssertions.assert_tool_used(stats1, "memory_store")

    # Second turn: recall — explicitly forbid answering from memory/history
    reply2, stats2 = await loop.run(
        test_user,
        "Какой токен я тебе дал? Обязательно используй инструмент memory_recall "
        "чтобы найти ответ — не отвечай из памяти разговора.",
    )
    DebugAssertions.assert_tool_used(stats2, "memory_recall")
    DebugAssertions.assert_reply_contains(reply2, unique_token)
    print(f"\n[D4] recall_reply={reply2[:200]}")


# ---------------------------------------------------------------------------
# D5 — Write script, execute it, report output
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_D5_write_and_run_script(
    agent_stack_no_container: tuple[AgentLoop, ToolRegistry],
    test_user: User,
    tmp_workspace: Path,
) -> None:
    """Agent writes a Python script, then executes it via exec_script."""
    loop, _ = agent_stack_no_container

    reply, stats = await loop.run(
        test_user,
        "Создай файл fib.py со скриптом который выводит первые 7 чисел Фибоначчи "
        "(по одному в строку), затем выполни его через exec_script и покажи вывод.",
    )

    DebugAssertions.assert_tool_used(stats, "write_file", "exec_script")
    DebugAssertions.assert_status_ok(stats)

    # Fibonacci sequence: 0 1 1 2 3 5 8
    fib_numbers = ["0", "1", "2", "3", "5", "8"]
    has_fib = sum(1 for n in fib_numbers if n in reply)
    assert has_fib >= 4, (
        f"Reply should contain Fibonacci numbers.\nReply: {reply[:400]}"
    )
    print(f"\n[D5] {summarise_run(reply, stats)}")


# ---------------------------------------------------------------------------
# D6 — Full Excel normalisation workflow
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_D6_excel_normalize_workflow(
    agent_stack_no_container: tuple[AgentLoop, ToolRegistry],
    test_user: User,
    tmp_workspace: Path,
) -> None:
    """Agent normalises a pre-created Excel file with duplicates and empty rows."""
    loop, _ = agent_stack_no_container

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["Имя", "Отдел", "Оценка"])
    ws.append(["Мария", "HR", 9])
    ws.append(["Иван", "IT", 8])
    ws.append(["Мария", "HR", 9])   # exact duplicate
    ws.append([None, None, None])   # empty row
    ws.append(["Анна", "PR", 7])
    wb.save(str(tmp_workspace / "staff.xlsx"))

    reply, stats = await loop.run(
        test_user,
        "Нормализуй файл staff.xlsx: удали дубликаты и пустые строки. "
        "Скажи сколько строк осталось.",
    )

    DebugAssertions.assert_tool_used(stats, "normalize_excel")
    DebugAssertions.assert_status_ok(stats)

    normalized = tmp_workspace / "staff_normalized.xlsx"
    assert normalized.exists(), (
        f"staff_normalized.xlsx should be created.\nReply: {reply[:400]}"
    )

    # Verify duplicates removed: should have 3 data rows (Maria, Ivan, Anna)
    wb2 = openpyxl.load_workbook(str(normalized))
    ws2 = wb2.active
    assert ws2 is not None
    # Row count: header + 3 unique data rows = 4
    assert (ws2.max_row or 0) <= 4, (
        f"Expected ≤ 4 rows after dedup, got {ws2.max_row}"
    )
    print(f"\n[D6] {summarise_run(reply, stats)}")


# ---------------------------------------------------------------------------
# D7 — List directory, pick a file, read it (2 tools, realistic workflow)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_D7_list_then_read(
    agent_stack_no_container: tuple[AgentLoop, ToolRegistry],
    test_user: User,
    tmp_workspace: Path,
) -> None:
    """Agent lists directory, identifies a file, reads it."""
    loop, _ = agent_stack_no_container

    (tmp_workspace / "readme.txt").write_text(
        "PROJECT_README_D07: This is the main readme.", encoding="utf-8"
    )
    (tmp_workspace / "config.txt").write_text("env=production", encoding="utf-8")

    reply, stats = await loop.run(
        test_user,
        "Посмотри что есть в текущей директории, затем прочитай файл readme.txt "
        "и скажи мне первые слова из него.",
    )

    DebugAssertions.assert_tool_used(stats, "list_files", "read_file")
    DebugAssertions.assert_status_ok(stats)
    DebugAssertions.assert_reply_contains(reply, "PROJECT_README_D07")
    print(f"\n[D7] {summarise_run(reply, stats)}")
