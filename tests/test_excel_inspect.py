"""Tests for ExcelInspectTool."""

from __future__ import annotations

import json
from pathlib import Path

import pytest

from corpclaw_lite.extensions.tools.builtin.excel_inspect import ExcelInspectTool


@pytest.fixture
def tool() -> ExcelInspectTool:
    return ExcelInspectTool()


def _create_simple_xlsx(path: Path) -> Path:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.cell(row=1, column=1, value="Name")
    ws.cell(row=1, column=2, value="Age")
    ws.cell(row=2, column=1, value="Alice")
    ws.cell(row=2, column=2, value=30)
    ws.cell(row=3, column=1, value="Bob")
    ws.cell(row=3, column=2, value=25)
    wb.save(str(path))
    return path


def _create_xlsx_with_merged_and_colors(path: Path) -> Path:
    import openpyxl
    from openpyxl.styles import PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    ws.merge_cells("A1:B1")
    ws["A1"] = "Merged Header"
    ws["A2"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws["A2"] = "Yellow cell"
    ws["B2"] = "Normal cell"
    ws["A3"] = "Row 3"
    ws["B3"] = "More data"
    wb.save(str(path))
    return path


class TestExcelInspectTool:
    @pytest.mark.asyncio
    async def test_inspect_xlsx_summary(
        self, tool: ExcelInspectTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        _create_simple_xlsx(tmp_path / "data.xlsx")

        result = await tool.execute(path="data.xlsx", detail="summary")
        assert "File: data.xlsx" in result
        assert "Sheets: 1" in result
        assert "Data" in result
        assert "3 rows x 2 cols" in result

    @pytest.mark.asyncio
    async def test_inspect_xlsx_full(
        self, tool: ExcelInspectTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        _create_xlsx_with_merged_and_colors(tmp_path / "merged.xlsx")

        result = await tool.execute(path="merged.xlsx", detail="full")
        assert "Merged Header" in result
        assert "A1:B1" in result
        assert "Preview" in result
        assert "Row 1:" in result

    @pytest.mark.asyncio
    async def test_inspect_xlsx_color_groups(
        self, tool: ExcelInspectTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        _create_xlsx_with_merged_and_colors(tmp_path / "colors.xlsx")

        result = await tool.execute(path="colors.xlsx", detail="full")
        assert "Color groups" in result
        # Compact summary format: "N colors detected (#hex: count, ...)"
        assert "FFFF00" in result
        assert "1 colors" in result
        # No cell coordinate lists in compact format
        assert "R2C1" not in result

    @pytest.mark.asyncio
    async def test_full_xlsx_inspection_size_limit(
        self, tool: ExcelInspectTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        _create_simple_xlsx(tmp_path / "large.xlsx")
        monkeypatch.setattr(
            "corpclaw_lite.extensions.tools.builtin.excel_inspect._MAX_FULL_XLSX_BYTES",
            1,
        )

        result = await tool.execute(path="large.xlsx", detail="full")

        assert "Error" in result
        assert "too large" in result

    @pytest.mark.asyncio
    async def test_inspect_csv(
        self, tool: ExcelInspectTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        (tmp_path / "data.csv").write_text(
            "name,city\nAlice,Moscow\nBob,SPb\nCarol,Kazan\n", encoding="utf-8"
        )

        result = await tool.execute(path="data.csv")
        assert "File: data.csv" in result
        assert "Format: CSV" in result
        assert "Columns (2): name, city" in result
        assert "Data rows: 3" in result
        assert "Preview" in result
        assert "Alice" in result

    @pytest.mark.asyncio
    async def test_inspect_large_csv_wording(
        self, tool: ExcelInspectTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Large CSV row count uses a clear lower-bound phrasing, not 'at least'."""
        monkeypatch.chdir(tmp_path)
        (tmp_path / "big.csv").write_text("a,b\n1,2\n3,4\n", encoding="utf-8")
        # Force the large-file branch: file is tiny, so lower both thresholds.
        monkeypatch.setattr(
            "corpclaw_lite.extensions.tools.builtin.excel_inspect._MAX_CSV_COUNT_BYTES",
            1,
        )
        monkeypatch.setattr(
            "corpclaw_lite.extensions.tools.builtin.excel_inspect._MAX_CSV_COUNT_ROWS",
            1,
        )

        result = await tool.execute(path="big.csv")
        assert "\u2265" in result  # ≥
        assert "skipped for performance" in result
        assert "at least" not in result

    @pytest.mark.asyncio
    async def test_inspect_json_array(
        self, tool: ExcelInspectTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        data = [{"x": 1, "y": "a"}, {"x": 2, "y": "b"}, {"x": 3, "y": "c"}]
        (tmp_path / "data.json").write_text(json.dumps(data), encoding="utf-8")

        result = await tool.execute(path="data.json")
        assert "File: data.json" in result
        assert "Format: JSON" in result
        assert "array of 3 items" in result
        assert "Columns" in result
        assert "Item 1:" in result

    @pytest.mark.asyncio
    async def test_inspect_json_object(
        self, tool: ExcelInspectTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        data = {"name": "Alice", "age": 30, "city": "Moscow"}
        (tmp_path / "obj.json").write_text(json.dumps(data), encoding="utf-8")

        result = await tool.execute(path="obj.json")
        assert "Format: JSON" in result
        assert "object with 3 keys" in result
        assert "name" in result
        assert "age" in result

    @pytest.mark.asyncio
    async def test_inspect_nonexistent_file(self, tool: ExcelInspectTool) -> None:
        result = await tool.execute(path="nonexistent.xlsx")
        assert "Error" in result
        assert "not found" in result.lower() or "File not found" in result

    @pytest.mark.asyncio
    async def test_inspect_unsupported_format(
        self, tool: ExcelInspectTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        (tmp_path / "data.txt").write_text("hello world", encoding="utf-8")

        result = await tool.execute(path="data.txt")
        assert "Error" in result
        assert "Unsupported" in result

    @pytest.mark.asyncio
    async def test_inspect_missing_path(self, tool: ExcelInspectTool) -> None:
        result = await tool.execute()
        assert "Error" in result
        assert "path" in result.lower()
