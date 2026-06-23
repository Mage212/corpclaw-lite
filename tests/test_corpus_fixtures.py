"""Tests for the xlsx corpus fixture generator (B-060)."""

from __future__ import annotations

from pathlib import Path

import pytest

from corpclaw_lite.eval.corpus_fixtures import (
    SUPPORTED_GENERATORS,
    generate_workbook,
    is_supported,
)


def test_messy_headers_generates_valid_xlsx(tmp_path: Path) -> None:
    """The messy_headers generator produces a valid xlsx with messy headers."""
    dest = tmp_path / "messy.xlsx"
    generate_workbook("messy_headers", dest)
    assert dest.exists()
    assert dest.stat().st_size > 0

    from openpyxl import load_workbook

    wb = load_workbook(dest)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    # Messy: extra spaces, mixed case — exactly what normalize_excel should fix.
    assert "  Name  " in headers
    assert "DEPARTMENT" in headers
    assert "salary  " in headers
    assert ws.max_row >= 2  # at least one data row


def test_unknown_generator_raises(tmp_path: Path) -> None:
    """An unknown generator id raises ValueError with the supported list."""
    with pytest.raises(ValueError, match="Unknown workbook generator"):
        generate_workbook("nonexistent", tmp_path / "out.xlsx")


def test_is_supported() -> None:
    assert is_supported("messy_headers")
    assert not is_supported("nonexistent")
    assert "messy_headers" in SUPPORTED_GENERATORS
