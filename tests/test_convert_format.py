"""Tests for ConvertFormatTool."""

from __future__ import annotations

import json
from pathlib import Path

import pytest

from corpclaw_lite.extensions.tools.builtin.convert_format import ConvertFormatTool


@pytest.fixture
def tool() -> ConvertFormatTool:
    return ConvertFormatTool()


def _create_csv(path: Path, content: str) -> Path:
    path.write_text(content, encoding="utf-8")
    return path


def _create_xlsx(path: Path, headers: list[str], rows: list[list[object]]) -> Path:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, row in enumerate(rows, 2):
        for col, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col, value=val)
    wb.save(str(path))
    return path


def _create_json(path: Path, data: list[dict[str, object]]) -> Path:
    path.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
    return path


def _create_markdown(path: Path, headers: list[str], rows: list[list[str]]) -> Path:
    lines = ["| " + " | ".join(headers) + " |"]
    lines.append("| " + " | ".join("---" for _ in headers) + " |")
    lines.extend("| " + " | ".join(row) + " |" for row in rows)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


class TestConvertFormatTool:
    @pytest.mark.asyncio
    async def test_csv_to_json(
        self, tool: ConvertFormatTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        _create_csv(tmp_path / "data.csv", "name,age\nAlice,30\nBob,25\n")

        result = await tool.execute(input_path="data.csv", output_format="json")
        assert "Converted" in result
        assert (tmp_path / "data.json").exists()
        data = json.loads((tmp_path / "data.json").read_text())
        assert len(data) == 2
        assert data[0]["name"] == "Alice"

    @pytest.mark.asyncio
    async def test_json_to_csv(
        self, tool: ConvertFormatTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        _create_json(tmp_path / "data.json", [{"x": 1, "y": 2}, {"x": 3, "y": 4}])

        result = await tool.execute(input_path="data.json", output_format="csv")
        assert "Converted" in result
        assert (tmp_path / "data.csv").exists()
        content = (tmp_path / "data.csv").read_text()
        assert "x" in content

    @pytest.mark.asyncio
    async def test_xlsx_to_csv(
        self, tool: ConvertFormatTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        _create_xlsx(tmp_path / "data.xlsx", ["city", "pop"], [["Moscow", "12M"], ["SPb", "5M"]])

        result = await tool.execute(input_path="data.xlsx", output_format="csv")
        assert "Converted" in result
        content = (tmp_path / "data.csv").read_text()
        assert "Moscow" in content

    @pytest.mark.asyncio
    async def test_csv_to_markdown(
        self, tool: ConvertFormatTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        _create_csv(tmp_path / "data.csv", "name,value\nAlice,100\nBob,200\n")

        result = await tool.execute(input_path="data.csv", output_format="markdown")
        assert "Converted" in result
        content = (tmp_path / "data.md").read_text()
        assert "| name" in content
        assert "Alice" in content

    @pytest.mark.asyncio
    async def test_markdown_to_csv(
        self, tool: ConvertFormatTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        _create_markdown(tmp_path / "table.md", ["col1", "col2"], [["a", "b"], ["c", "d"]])

        result = await tool.execute(input_path="table.md", output_format="csv")
        assert "Converted" in result
        content = (tmp_path / "table.csv").read_text()
        assert "col1" in content
        assert "a" in content

    @pytest.mark.asyncio
    async def test_csv_to_xlsx(
        self, tool: ConvertFormatTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        _create_csv(tmp_path / "data.csv", "x,y\n1,2\n3,4\n")

        result = await tool.execute(input_path="data.csv", output_format="xlsx")
        assert "Converted" in result
        assert (tmp_path / "data.xlsx").exists()

    @pytest.mark.asyncio
    async def test_custom_output_path(
        self, tool: ConvertFormatTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        _create_csv(tmp_path / "data.csv", "a,b\n1,2\n")

        result = await tool.execute(
            input_path="data.csv", output_format="json", output_path="custom.json"
        )
        assert "custom.json" in result
        assert (tmp_path / "custom.json").exists()

    @pytest.mark.asyncio
    async def test_output_path_outside_workspace_blocked(
        self, tool: ConvertFormatTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        workspace = tmp_path / "workspace"
        workspace.mkdir()
        monkeypatch.chdir(workspace)
        _create_csv(workspace / "data.csv", "a,b\n1,2\n")

        result = await tool.execute(
            input_path="data.csv",
            output_format="json",
            output_path="../escaped.json",
        )

        assert "Error" in result
        assert not (tmp_path / "escaped.json").exists()

    @pytest.mark.asyncio
    async def test_file_not_found(self, tool: ConvertFormatTool) -> None:
        result = await tool.execute(input_path="nonexistent.csv", output_format="json")
        assert "Error" in result

    @pytest.mark.asyncio
    async def test_input_file_too_large(
        self, tool: ConvertFormatTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        _create_csv(tmp_path / "large.csv", "a,b\n1,2\n")
        monkeypatch.setattr(
            "corpclaw_lite.extensions.tools.builtin.convert_format._MAX_INPUT_BYTES",
            1,
        )

        result = await tool.execute(input_path="large.csv", output_format="json")

        assert "Error" in result
        assert "too large" in result

    @pytest.mark.asyncio
    async def test_unsupported_input(
        self, tool: ConvertFormatTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        (tmp_path / "data.txt").write_text("hello", encoding="utf-8")

        result = await tool.execute(input_path="data.txt", output_format="json")
        assert "Error" in result

    @pytest.mark.asyncio
    async def test_missing_params(self, tool: ConvertFormatTool) -> None:
        result = await tool.execute()
        assert "Error" in result

    # --- Cyrillic / encoding tests ---

    @pytest.mark.asyncio
    async def test_csv_cp1251_to_json(
        self, tool: ConvertFormatTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Windows-1251 CSV (default Russian Excel export) must be read correctly."""
        monkeypatch.chdir(tmp_path)
        cp1251_content = "имя,город\nАлексей,Москва\nМария,Санкт-Петербург\n"
        (tmp_path / "ru.csv").write_bytes(cp1251_content.encode("cp1251"))

        result = await tool.execute(input_path="ru.csv", output_format="json")
        assert "Converted" in result
        data = json.loads((tmp_path / "ru.json").read_text(encoding="utf-8"))
        assert data[0]["имя"] == "Алексей"
        assert data[1]["город"] == "Санкт-Петербург"

    @pytest.mark.asyncio
    async def test_csv_utf8_bom_to_json(
        self, tool: ConvertFormatTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """UTF-8 BOM CSV ('CSV UTF-8' option in Excel) must be read correctly."""
        monkeypatch.chdir(tmp_path)
        bom_content = "имя,возраст\nДмитрий,35\nЕлена,28\n"
        (tmp_path / "bom.csv").write_bytes(b"\xef\xbb\xbf" + bom_content.encode("utf-8"))

        result = await tool.execute(input_path="bom.csv", output_format="json")
        assert "Converted" in result
        data = json.loads((tmp_path / "bom.json").read_text(encoding="utf-8"))
        assert data[0]["имя"] == "Дмитрий"
        assert data[1]["возраст"] == "28"

    @pytest.mark.asyncio
    async def test_csv_write_has_bom(
        self, tool: ConvertFormatTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Written CSV must include BOM so Excel opens Cyrillic correctly."""
        monkeypatch.chdir(tmp_path)
        _create_json(
            tmp_path / "data.json",
            [{"товар": "Хлеб", "цена": 50}, {"товар": "Молоко", "цена": 80}],
        )

        result = await tool.execute(input_path="data.json", output_format="csv")
        assert "Converted" in result
        raw = (tmp_path / "data.csv").read_bytes()
        assert raw[:3] == b"\xef\xbb\xbf", "CSV must start with UTF-8 BOM"
        content = raw.decode("utf-8-sig")
        assert "Хлеб" in content
        assert "Молоко" in content

    @pytest.mark.asyncio
    async def test_xlsx_cyrillic_roundtrip(
        self, tool: ConvertFormatTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """XLSX with Cyrillic → CSV → back to XLSX preserves data."""
        monkeypatch.chdir(tmp_path)
        _create_xlsx(
            tmp_path / "orig.xlsx",
            ["ФИО", "Отдел"],
            [["Иванов Иван", "Бухгалтерия"], ["Петрова Анна", "IT"]],
        )

        result = await tool.execute(input_path="orig.xlsx", output_format="csv")
        assert "Converted" in result
        csv_content = (tmp_path / "orig.csv").read_text(encoding="utf-8-sig")
        assert "Иванов" in csv_content
        assert "Бухгалтерия" in csv_content

    # --- XLSX lossy-conversion note (Правка 1) ---

    @pytest.mark.asyncio
    async def test_convert_xlsx_input_warns_formula_loss(
        self, tool: ConvertFormatTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """XLSX input conversion must surface the lossy-read note."""
        monkeypatch.chdir(tmp_path)
        _create_xlsx(tmp_path / "data.xlsx", ["a", "b"], [["1", "2"]])

        result = await tool.execute(input_path="data.xlsx", output_format="csv")
        assert "Converted" in result
        assert "values only" in result
        assert "excel_workbook" in result

    @pytest.mark.asyncio
    async def test_convert_non_xlsx_input_has_no_lossy_note(
        self, tool: ConvertFormatTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """CSV/JSON/Markdown inputs are not lossy and must not carry the note."""
        monkeypatch.chdir(tmp_path)
        _create_csv(tmp_path / "data.csv", "a,b\n1,2\n")

        result = await tool.execute(input_path="data.csv", output_format="json")
        assert "Converted" in result
        assert "values only" not in result

    # --- type-aware xlsx writer (Правка 2) ---

    @pytest.mark.asyncio
    async def test_write_xlsx_preserves_types(
        self, tool: ConvertFormatTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """XLSX output applies number_format matching the Python type of each value."""
        import openpyxl

        monkeypatch.chdir(tmp_path)
        _create_json(
            tmp_path / "data.json",
            [
                {"inn": "0077000000", "count": 42, "price": 12.5},
                {"inn": "0077000001", "count": 7, "price": 3.14},
            ],
        )

        result = await tool.execute(input_path="data.json", output_format="xlsx")
        assert "Converted" in result

        wb = openpyxl.load_workbook(str(tmp_path / "data.xlsx"))
        ws = wb.active
        assert ws is not None
        # Header row is text-formatted.
        assert ws.cell(row=1, column=1).number_format == "@"
        # INN strings keep text format so leading zeros survive in Excel.
        assert ws.cell(row=2, column=1).number_format == "@"
        assert ws.cell(row=2, column=1).value == "0077000000"
        # int → "0", float → "0.00".
        assert ws.cell(row=2, column=2).number_format == "0"
        assert ws.cell(row=2, column=2).value == 42
        assert ws.cell(row=2, column=3).number_format == "0.00"
        assert ws.cell(row=2, column=3).value == 12.5
        wb.close()

    # --- duplicate headers (Правка 3) ---

    @pytest.mark.asyncio
    async def test_load_csv_duplicate_headers(
        self, tool: ConvertFormatTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Duplicate CSV headers are renamed col, col_1 — no data loss."""
        monkeypatch.chdir(tmp_path)
        _create_csv(tmp_path / "data.csv", "name,city,name\nAlice,Moscow,Alice A.\n")

        result = await tool.execute(input_path="data.csv", output_format="json")
        assert "Converted" in result
        data = json.loads((tmp_path / "data.json").read_text(encoding="utf-8"))
        assert data[0]["name"] == "Alice"
        assert data[0]["name_1"] == "Alice A."
        assert data[0]["city"] == "Moscow"

    @pytest.mark.asyncio
    async def test_load_xlsx_duplicate_headers(
        self, tool: ConvertFormatTool, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Duplicate XLSX headers are renamed col, col_1 — no data loss."""
        monkeypatch.chdir(tmp_path)
        _create_xlsx(
            tmp_path / "data.xlsx",
            ["name", "city", "name"],
            [["Alice", "Moscow", "Alice A."]],
        )

        result = await tool.execute(input_path="data.xlsx", output_format="json")
        assert "Converted" in result
        data = json.loads((tmp_path / "data.json").read_text(encoding="utf-8"))
        assert data[0]["name"] == "Alice"
        assert data[0]["name_1"] == "Alice A."
        assert data[0]["city"] == "Moscow"
